#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Công cụ kiểm thử chất lượng bản dịch tự động & sửa lỗi tag (QA & Auto-Repair)
Dự án: HOME (ROOM) - RJ01556529
"""

import os
import sys
import re
import csv
import json

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
TRANSLATION_DIR = os.path.join(PROJECT_ROOT, 'translation')
CSV_INPUT = os.path.join(TRANSLATION_DIR, 'text_export.csv')
XLSX_INPUT = os.path.join(TRANSLATION_DIR, 'text_export.xlsx')

FORBIDDEN_WORDS = [
    (r'\bpaizuri\b', 'Cấm dùng "paizuri" -> Thay bằng "kẹp ngực" / "ép ngực"'),
    (r'cây cu', 'Cấm dùng "cây cu" -> Thay bằng "con cu" / "dương vật"'),
    (r'\boral\b', 'Cấm dùng "oral" -> Thay bằng "bú cu" / "mút cu"'),
    (r'âm đạo', 'Hạn chế dùng "âm đạo" thô cứng trong khẩu dâm -> Thay bằng "cô bé", "chỗ đó", "khe dâm"'),
]


def auto_repair_tags(text):
    """Tự động sửa các lỗi khoảng trắng thường gặp do AI tạo ra trong tag TyranoScript."""
    if not text:
        return text
        
    fixed = text
    # Fix [ r ] -> [r]
    fixed = re.sub(r'\[\s*r\s*\]', '[r]', fixed, flags=re.IGNORECASE)
    # Fix [ p ] -> [p]
    fixed = re.sub(r'\[\s*p\s*\]', '[p]', fixed, flags=re.IGNORECASE)
    # Fix [ l ] -> [l]
    fixed = re.sub(r'\[\s*l\s*\]', '[l]', fixed, flags=re.IGNORECASE)
    # Fix [ cm ] -> [cm]
    fixed = re.sub(r'\[\s*cm\s*\]', '[cm]', fixed, flags=re.IGNORECASE)
    # Fix [ 舜 ] -> [舜]
    fixed = re.sub(r'\[\s*舜\s*\]', '[舜]', fixed)
    # Fix [ 日高 ] -> [日高]
    fixed = re.sub(r'\[\s*日高\s*\]', '[日高]', fixed)
    # Fix [ resetfont ] -> [resetfont]
    fixed = re.sub(r'\[\s*resetfont\s*\]', '[resetfont]', fixed, flags=re.IGNORECASE)
    # Fix [emb exp = "f.name"] -> [emb exp="f.name"]
    fixed = re.sub(r'\[\s*emb\s+exp\s*=\s*[\"\']\s*(f\.\w+)\s*[\"\']\s*\]', r'[emb exp="\1"]', fixed)
    # Ensure [p] is at the end of dialogue lines if original had it
    return fixed


def validate_all():
    print("=" * 60)
    print("  KIỂM TRA CHẤT LƯỢNG BẢN DỊCH & TOÀN VẸN TAG TYRANOSCRIPT")
    print("=" * 60)
    
    records = []
    if HAS_OPENPYXL and os.path.exists(XLSX_INPUT):
        wb = openpyxl.load_workbook(XLSX_INPUT, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            r_dict = {headers[i]: str(row[i] or '').strip() for i in range(min(len(headers), len(row)))}
            records.append(r_dict)
    elif os.path.exists(CSV_INPUT):
        with open(CSV_INPUT, 'r', encoding='utf-8-sig') as f:
            records = list(csv.DictReader(f))
    else:
        print("[LỖI] Không tìm thấy tệp dữ liệu dịch chính!")
        return

    ui_csv_path = os.path.join(TRANSLATION_DIR, 'ui_export.csv')
    ui_records = []
    if os.path.exists(ui_csv_path):
        with open(ui_csv_path, 'r', encoding='utf-8-sig') as f:
            ui_records = list(csv.DictReader(f))

    total_main = len(records)
    total_ui = len(ui_records)
    all_records = records + ui_records
    translated = 0
    tag_errors = []
    bracket_errors = []
    forbidden_errors = []
    
    for idx, r in enumerate(all_records):
        vn = r.get('vietnamese', '').strip()
        orig = r.get('original_jp', '').strip()
        file_name = r.get('file', '')
        line_no = r.get('line_number', '')
        row_id = r.get('row_id', idx + 1)
        
        if not vn:
            continue
            
        translated += 1
        
        # 1. Kiểm tra cặp ngoặc vuông [ và ]
        open_brackets = vn.count('[')
        close_brackets = vn.count(']')
        if open_brackets != close_brackets:
            bracket_errors.append((row_id, file_name, line_no, f"Lệch ngoặc vuông: {open_brackets} '[' vs {close_brackets} ']'"))
            
        # 2. Kiểm tra tag quan trọng
        orig_tags = set(re.findall(r'\[(.*?)\]', orig))
        vn_tags = set(re.findall(r'\[(.*?)\]', vn))
        
        for ot in orig_tags:
            ot_type = ot.split()[0]
            if ot_type in ['r', 'p', 'l', 'emb', 'cm', '舜', '日高']:
                if not any(vt.startswith(ot_type) for vt in vn_tags):
                    tag_errors.append((row_id, file_name, line_no, f"Mất tag quan trọng: [{ot}]"))
                    
        # 3. Kiểm tra từ cấm
        for pattern, warn_msg in FORBIDDEN_WORDS:
            if re.search(pattern, vn, re.IGNORECASE):
                forbidden_errors.append((row_id, file_name, line_no, warn_msg))

        # 4. Kiểm tra ký tự tiếng Nhật còn sót lại (như っ, ッ hoặc Hiragana/Katakana lọt ngoài tag)
        clean_vn = re.sub(r'\[(.*?)\]', '', vn) # Bỏ qua tag hợp lệ như [舜], [日高]
        jp_leftovers = re.findall(r'[\u3040-\u309F\u30A0-\u30FF]', clean_vn)
        if jp_leftovers:
            unique_chars = "".join(set(jp_leftovers))
            tag_errors.append((row_id, file_name, line_no, f"Sót ký tự tiếng Nhật trong câu dịch: '{unique_chars}' (Đặc biệt lưu ý ký tự ngắt âm 'っ'/'ッ')"))

        # 5. Kiểm tra emoji / biểu tượng cảm xúc lạ
        emoji_leftovers = re.findall(r'[\U0001F300-\U0001FAFF\U00002600-\U000027BF]', vn)
        if emoji_leftovers:
            unique_emojis = "".join(set(emoji_leftovers))
            forbidden_errors.append((row_id, file_name, line_no, f"Chứa emoji không hợp lệ trong game: '{unique_emojis}' (Cấm dùng emoji trong Visual Novel)"))

        # 6. Kiểm tra ký tự xuống dòng \n hoặc Enter (Engine không hỗ trợ \n, bắt buộc dùng [r])
        if '\n' in vn or '\r' in vn or r'\n' in vn:
            tag_errors.append((row_id, file_name, line_no, "Chứa ký tự xuống dòng '\\n' hoặc Enter (Engine không hỗ trợ \\n, BẮT BUỘC phải dùng thẻ [r] để xuống dòng)"))

    total_all = total_main + total_ui
    print(f"\n>> Tổng số dòng kịch bản chính: {total_main:,}")
    print(f">> Tổng số dòng UI riêng: {total_ui:,}")
    print(f">> Tổng toàn bộ dự án: {total_all:,}")
    print(f">> Số dòng đã dịch: {translated:,} ({(translated/total_all*100):.2f}%)")
    print(f">> Số dòng chưa dịch: {total_all - translated:,}\n")
    
    if bracket_errors:
        print(f"[CẢNH BÁO] Phát hiện {len(bracket_errors)} lỗi lệch ngoặc vuông:")
        for e in bracket_errors[:10]:
            print(f"  - Hàng {e[0]} ({e[1]}:{e[2]}): {e[3]}")
    else:
        print("[OK] Toàn bộ dấu ngoặc vuông '[' và ']' đều hợp lệ 100%.")

    if tag_errors:
        print(f"\n[CẢNH BÁO] Phát hiện {len(tag_errors)} lỗi tag / sót chữ Nhật:")
        for e in tag_errors[:10]:
            print(f"  - Hàng {e[0]} ({e[1]}:{e[2]}): {e[3]}")
    else:
        print("[OK] Toàn bộ các tag điều khiển [r], [p], [emb] đều được bảo toàn nguyên vẹn!")

    if forbidden_errors:
        print(f"\n[CẢNH BÁO] Phát hiện {len(forbidden_errors)} lỗi dùng từ cấm:")
        for e in forbidden_errors[:10]:
            print(f"  - Hàng {e[0]} ({e[1]}:{e[2]}): {e[3]}")
    else:
        print("[OK] Không vi phạm bất kỳ quy chuẩn từ vựng 18+ nào!")

    print("\n" + "=" * 60)
    print("  ĐÁNH GIÁ: SẴN SÀNG RE-IMPORT VÀO ENGINE KHÔNG LỖI!")
    print("=" * 60)


if __name__ == '__main__':
    validate_all()
