#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script tái nhập (Re-import) bản dịch từ CSV/XLSX vào tệp kịch bản .ks của TyranoScript
Dự án: HOME - RJ01556529
"""

import os
import sys
import re
import csv
import json
from collections import defaultdict

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SCENARIO_SRC_DIR = os.path.join(PROJECT_ROOT, 'extracted_scripts', 'data', 'scenario')
TRANSLATION_DIR = os.path.join(PROJECT_ROOT, 'translation')
CSV_INPUT = os.path.join(TRANSLATION_DIR, 'text_export.csv')
XLSX_INPUT = os.path.join(TRANSLATION_DIR, 'text_export.xlsx')
OUTPUT_PATCH_DIR = os.path.join(PROJECT_ROOT, 'patch', 'data', 'scenario')


def load_translations():
    """Đọc dữ liệu dịch từ XLSX/CSV chính và UI CSV riêng."""
    records = []
    ui_records = []
    
    # 1. Đọc file chính: Ưu tiên file Excel XLSX nếu có
    if HAS_OPENPYXL and os.path.exists(XLSX_INPUT):
        print(f"Đang đọc dữ liệu chính từ file Excel: {XLSX_INPUT}...")
        wb = openpyxl.load_workbook(XLSX_INPUT, data_only=True)
        ws = wb.active
        headers = [str(cell.value or '').strip() for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for col_idx, h in enumerate(headers):
                if col_idx < len(row):
                    val = row[col_idx]
                    row_dict[h] = str(val if val is not None else '').strip()
                else:
                    row_dict[h] = ''
            if row_dict.get('file') and row_dict.get('line_number'):
                records.append(row_dict)
                
    elif os.path.exists(CSV_INPUT):
        print(f"Đang đọc dữ liệu chính từ file CSV: {CSV_INPUT}...")
        with open(CSV_INPUT, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.DictReader(f)
            for r in reader:
                records.append({k: str(v).strip() for k, v in r.items()})
    else:
        print(f"[WARN] Không tìm thấy file dịch chính {XLSX_INPUT} hoặc {CSV_INPUT}")

    # 2. Đọc file UI riêng nếu có (ui_export.csv)
    ui_csv_path = os.path.join(TRANSLATION_DIR, 'ui_export.csv')
    if os.path.exists(ui_csv_path):
        print(f"Đang đọc thêm dữ liệu UI từ: {ui_csv_path}...")
        with open(ui_csv_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.DictReader(f)
            for r in reader:
                ui_records.append({k: str(v).strip() for k, v in r.items()})

    combined = records + ui_records
    print(f"Đã nạp tổng cộng {len(combined)} bản ghi ({len(records)} kịch bản + {len(ui_records)} UI).")
    return combined


def auto_guard_leading_tags(orig_text, trans_text):
    """Tự động bảo hộ và khôi phục các thẻ điều khiển đầu dòng (như [chara_mod ...], [playse ...], [image ...])
    nếu AI vô tình làm mất hoặc xóa đi.
    """
    if not trans_text:
        return trans_text
        
    leading_tags_match = re.match(r'^((?:\[(?:chara_mod|chara_show|chara_hide|image|playse|stopse|anim|bg|mask|fadein|quake)[^\]]*\]\s*)+)', orig_text)
    if leading_tags_match:
        leading_tags = leading_tags_match.group(1).strip()
        if not trans_text.strip().startswith(leading_tags):
            cleaned_trans = re.sub(r'\[(?:chara_mod|chara_show|chara_hide|image|playse|stopse|anim|bg)[^\]]*\]\s*', '', trans_text).strip()
            return f"{leading_tags} {cleaned_trans}" if not cleaned_trans.startswith('「') and not cleaned_trans.startswith('凪') else f"{leading_tags}{cleaned_trans}"
            
    return trans_text


def validate_tag_parity(orig_text, trans_text):
    """Kiểm tra xem bản dịch có làm mất các tag TyranoScript quan trọng không."""
    orig_tags = re.findall(r'\[(.*?)\]', orig_text)
    trans_tags = re.findall(r'\[(.*?)\]', trans_text)
    
    # Check essential control tags
    for tag in orig_tags:
        # Ignore font styling or macros if slightly altered, but enforce [r], [p], [l], [emb]
        tag_type = tag.split()[0] if tag else ''
        if tag_type in ['r', 'p', 'l', 'emb', 'cm']:
            if not any(t.startswith(tag_type) for t in trans_tags):
                return False, f"Thiếu tag quan trọng: [{tag}]"
                
    return True, ""


def reimport_and_patch(records, output_dir=None):
    """
    Nạp dữ liệu đã dịch và thay thế chính xác theo chỉ số dòng vào file .ks
    """
    if output_dir is None:
        output_dir = OUTPUT_PATCH_DIR
    os.makedirs(output_dir, exist_ok=True)
    
    # Group translations by filename and line_number
    file_map = defaultdict(dict)
    translated_count = 0
    warning_count = 0
    
    for r in records:
        f_name = r.get('file', '').strip()
        try:
            line_no = int(r.get('line_number', 0))
        except ValueError:
            continue
            
        vn_text = r.get('vietnamese', '').strip()
        # Remove leading quote added to prevent formula interpretation
        if vn_text.startswith("'"):
            vn_text = vn_text[1:]
            
        if vn_text:
            translated_count += 1
            file_map[f_name][line_no] = {
                'vietnamese': vn_text,
                'original_jp': r.get('original_jp', ''),
                'entry_type': r.get('entry_type', ''),
                'speaker_raw': r.get('speaker_raw', '')
            }

    print(f"Tổng số câu đã được dịch: {translated_count}/{len(records)} câu.")

    # Process all files in source
    source_files = [f for f in os.listdir(SCENARIO_SRC_DIR) if f.endswith('.ks')]
    patched_files_count = 0
    
    for f_name in source_files:
        src_path = os.path.join(SCENARIO_SRC_DIR, f_name)
        dst_path = os.path.join(output_dir, f_name)
        
        lines = open(src_path, 'r', encoding='utf-8', errors='replace').readlines()
        # Loại bỏ dấu ] thừa ở dòng 1 do TyranoBuilder xuất lỗi
        if lines and lines[0].strip() == ']':
            lines[0] = '\n'
        mod_lines = list(lines)
        file_translations = file_map.get(f_name, {})
        has_mod = False
        
        for line_no, trans_info in file_translations.items():
            line_idx = line_no - 1
            if line_idx >= len(mod_lines):
                continue
                
            orig_raw_line = mod_lines[line_idx]
            clean_orig_raw = orig_raw_line.strip()
            vn_text = trans_info['vietnamese']
            orig_jp = trans_info['original_jp']
            entry_type = trans_info['entry_type']
            
            # Phòng vệ: Không bao giờ ghi đè lên các thẻ điều khiển [if], [endif], [jump], [macro] nếu không khớp
            if any(clean_orig_raw.startswith(cmd) for cmd in ['[if', '[endif', '[jump', '[call', '[macro', '[iscript', '[endscript', '[s', '[cm', '@if', '@endif']):
                if clean_orig_raw != orig_jp.strip():
                    print(f"  [PHÒNG VỆ CODE] Bỏ qua ghi đè thẻ hệ thống tại {f_name}:{line_no} (Gốc: '{clean_orig_raw}')")
                    continue
            
            # Tự động bảo vệ và khôi phục các thẻ điều khiển [chara_mod ...] ở đầu câu
            vn_text = auto_guard_leading_tags(orig_jp, vn_text)

            # Tag validation
            is_valid, warn_msg = validate_tag_parity(orig_jp, vn_text)
            if not is_valid:
                warning_count += 1
                # print(f"[CẢNH BÁO] {f_name}:{line_no} -> {warn_msg}")

            # Apply replacement based on entry_type
            if entry_type == 'glink_choice':
                # Replace text="..." in glink
                mod_line = re.sub(r'text=[\"\'].*?[\"\']', f'text="{vn_text}"', orig_raw_line, count=1)
                mod_lines[line_idx] = mod_line
                has_mod = True
            elif entry_type == 'ptext_ui':
                # Replace text="..." in ptext
                mod_line = re.sub(r'text=[\"\'].*?[\"\']', f'text="{vn_text}"', orig_raw_line, count=1)
                mod_lines[line_idx] = mod_line
                has_mod = True
            else:
                # Standard dialogue / narration line replacement
                # Preserve leading / trailing whitespace of original line
                leading_ws = len(orig_raw_line) - len(orig_raw_line.lstrip())
                trailing_ws = '\n' if orig_raw_line.endswith('\n') else ''
                indent = orig_raw_line[:leading_ws]
                
                mod_lines[line_idx] = indent + vn_text + trailing_ws
                has_mod = True

        # Sửa chữa tự động các thẻ bị mất dấu đóng ] ở cuối tệp & Việt hóa thẻ tên #name
        SPEAKER_NAMES = {
            '#凪': '#Nagi',
            '#蕾': '#Tsubomi',
            '#凛子': '#Rinko',
            '#ガイド': '#Hướng dẫn',
        }
        clean_lines = []
        for line in mod_lines:
            s = line.strip()
            if s in SPEAKER_NAMES:
                clean_lines.append(SPEAKER_NAMES[s] + '\n')
                has_mod = True
            elif s == '[retur':
                clean_lines.append('[return]\n')
            elif s.startswith('[') and not s.startswith(';') and s.count('[') > s.count(']'):
                clean_lines.append(line.rstrip('\r\n') + ']\n')
            else:
                clean_lines.append(line)

        # Write output file
        with open(dst_path, 'w', encoding='utf-8') as out_f:
            out_f.writelines(clean_lines)
            
        if has_mod:
            patched_files_count += 1

    apply_engine_typo_fixes(output_dir)

    print(f"[OK] Đã xuất {len(source_files)} tệp .ks vào thư mục patch: {output_dir}")
    print(f"[OK] Số tệp có chứa bản dịch đã vá: {patched_files_count} tệp.")
    if warning_count > 0:
        print(f"[WARN] Số dòng có cảnh báo tag: {warning_count} dòng.")
    else:
        print("[OK] Không có bất kỳ lỗi cú pháp hoặc mất tag nào!")


def apply_engine_typo_fixes(target_dir):
    """Sửa chữa tự động các lỗi typo trong kịch bản gốc của nhà phát triển Nhật Bản."""
    fixes = {
        'character.ks': [
            (True, r'target=[\"\']\*nagi_date_(?:kirai|nigate|0)[\"\']', 'target="*nagi_date_hutuu"'),
        ],
        'EV_tousatuCG.ks': [
            (True, r'\[call\s+storage=[\"\']EV_tousatuCG\.ks[\"\']\s+target=[\"\']\*wasitu_denki[\"\']\]', '; [call storage="EV_tousatuCG.ks" target="*wasitu_denki"]'),
        ],
        'H_3P.ks': [
            (False, 'H_serihu_trinkotubomi.ks', 'H_serihu_rinkotubomi.ks'),
            (False, 'H_rinkotubomi.ks', 'H_serihu_rinkotubomi.ks'),
        ],
        'H_3P0.ks': [
            (False, 'H_rinkotubomi.ks', 'H_serihu_rinkotubomi.ks'),
        ],
        'H_3P_2.ks': [
            (False, 'H_serihu_trinkotubomi.ks', 'H_serihu_rinkotubomi.ks'),
            (False, 'H_rinkotubomi.ks', 'H_serihu_rinkotubomi.ks'),
            (False, '*3P_Dkiss_tubomi_hit', '*3P_2_Dkiss_tubomi1'),
            (False, '*3P_Dkiss_rinko_hit', '*3P_2_Dkiss_tubomi1'),
        ],
        'H_nagi_gauge.ks': [
            (False, '*zettyou7', '*zettyou6'),
            (False, '*zettyou8', '*zettyou6'),
            (False, '*zettyou9', '*zettyou6'),
            (False, '*zettyou10', '*zettyou6'),
        ],
        'H_rinko_gauge.ks': [
            (False, '*zettyou7', '*zettyou6'),
            (False, '*zettyou8', '*zettyou6'),
            (False, '*zettyou9', '*zettyou6'),
            (False, '*zettyou10', '*zettyou6'),
        ],
        'H_tubomi_gauge.ks': [
            (False, '*zettyou7', '*zettyou6'),
            (False, '*zettyou8', '*zettyou6'),
            (False, '*zettyou9', '*zettyou6'),
            (False, '*zettyou10', '*zettyou6'),
        ],
        'H_rinko_supiritasu.ks': [
            (False, '*supiritasu_tekoki_hit', '*hit'),
            (False, '*tekoki_hit', '*hit'),
        ],
        'H_tubomi_R2.ks': [
            (False, '*R2_anaruseme1_hit', '*hit'),
            (False, '*R2_anaruseme_hit', '*hit'),
        ],
        'komyu_hayato.ks': [
            (False, '*kaiwa_0kirai', '*kaiwa_1nigate'),
            (False, '*kaiwa_0', '*kaiwa_1nigate'),
        ],
        'komyu_nagi_kaeru.ks': [
            (False, 'deto_gohan.ks', 'EV_deto_dinner.ks'),
            (False, 'deto_gohanEV.ks', 'EV_deto_dinner.ks'),
            (True, r'storage=[\"\']character\.ks[\"\']\s+target=[\"\']\*(?:nagi|chara_nagi)[\"\']', 'storage="character.ks" target="*nagi_sotogi"'),
        ],
        'room_asa.ks': [
            (True, r'\[jump\s+storage=[\"\']room_asa\.ks[\"\']\s+target=[\"\']\*mission[\"\']\s*\]', '[s]'),
        ],
        'sansaku_massajiEV.ks': [
            (True, r'storage=[\"\']sansaku_massajiEV\.ks[\"\'].*?target=[\"\']\*kaeru[\"\']', 'storage="sansaku.ks" target="*end"'),
        ],
        'ui_base.ks': [
            (True, r'target=[\"\']\*ten_takai[\"\']', 'target="*ten_takai3"'),
            (True, r'target=[\"\']\*ten_hikui[\"\']', 'target="*ten_hikui1"'),
            (True, r'target=[\"\']\*ten_hutuu[\"\']', 'target="*ten_hutuu1"'),
        ],
        'sinnyu_item.ks': [
            (False, 'hint=媚薬プレミアム', 'hint="Thuốc kích dục cao cấp"'),
            (False, 'hint=媚薬クリーム', 'hint="Kem kích dục"'),
            (False, 'hint=媚薬', 'hint="Thuốc kích dục"'),
            (False, 'hint=睡眠薬', 'hint="Thuốc ngủ"'),
        ],
        'sinnyu_PC.ks': [
            (False, '<div id="slide_exit_btn">終了</div>', '<div id="slide_exit_btn">Thoát</div>'),
            (False, "face=\"sans-serif,'メイリオ'\"", 'face="NotoSansVN, sans-serif"'),
        ],
        'CG_tou_complete.ks': [
            (False, "face=\"serif,'游明朝'\"", 'face="NotoSansVN, serif"'),
        ],
        'mesi_jisui.ks': [
            (False, 'n][_tb_system_call storage=system/_mesi_jisui.ks]', '[_tb_system_call storage=system/_mesi_jisui.ks]'),
        ],
        'minigame.ks': [
            (False, '[/iscript]', '[endscript]'),
            (False, '\\"ヒット！スコア: \\"', '\\"Trúng đích! Điểm số: \\"'),
            (False, '\\"ミス！スコア: \\"', '\\"Trượt rồi! Điểm số: \\"'),
            (False, '\\"ゲーム終了！最終スコア: \\"', '\\"Kết thúc trò chơi! Điểm chung cuộc: \\"'),
            (False, 'text = "リトライ"', 'text = "Thử lại"'),
        ],
    }

    for fname, rule_list in fixes.items():
        fpath = os.path.join(target_dir, fname)
        if os.path.exists(fpath):
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            for is_regex, old_pat, new_str in rule_list:
                if is_regex:
                    content = re.sub(old_pat, new_str, content)
                else:
                    content = content.replace(old_pat, new_str)
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(content)


if __name__ == '__main__':
    records = load_translations()
    reimport_and_patch(records)
    print("\n>>> RE-IMPORT KIỂM THỬ HOÀN TẤT!")
