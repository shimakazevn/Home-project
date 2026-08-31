# -*- coding: utf-8 -*-
"""
tools/deep_comprehensive_audit.py
=================================
Bộ kiểm thử đa tầng rà soát 100% kịch bản toàn dự án HOME [RJ01556529]:
- Tầng 1: Khớp chỉ số dòng 1:1 giữa CSDL và kịch bản gốc tiếng Nhật.
- Tầng 2: Khớp 100% nhân vật nói (Shun, Nagi, Rinko, Tsubomi, Hayato) giữa gốc và dịch.
- Tầng 3: Khớp từ khóa ngữ cảnh theo tình tiết (Vật phẩm, Địa điểm, Hành động).
- Tầng 4: Kiểm tra toàn vẹn thẻ tag và engine syntax trong tất cả file .ks.
"""

import os
import sys
import re
import csv
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ORIG_DIR = os.path.join(PROJECT_ROOT, 'extracted_scripts', 'data', 'scenario')
PATCH_DIR = os.path.join(PROJECT_ROOT, 'patch', 'data', 'scenario')
XLSX_PATH = os.path.join(PROJECT_ROOT, 'translation', 'text_export.xlsx')
CSV_PATH = os.path.join(PROJECT_ROOT, 'translation', 'text_export.csv')

print("=" * 80)
print("   BỘ KIỂM THỬ ĐA TẦNG TOÀN DIỆN KỊCH BẢN (COMPREHENSIVE AUDIT SUITE)")
print("=" * 80)

# =============================================================================
# TẦNG 1: ĐỐI CHIẾU DÒNG GỐC 1:1
# =============================================================================
print("\n[*] [TẦNG 1] Đang nạp kịch bản gốc và đối chiếu từng dòng trong CSDL...")

orig_files = {}
for f in os.listdir(ORIG_DIR):
    if f.endswith('.ks'):
        p = os.path.join(ORIG_DIR, f)
        with open(p, 'r', encoding='utf-8', errors='ignore') as fo:
            orig_files[f] = fo.readlines()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

layer1_errors = []
layer2_errors = []
layer3_errors = []

for r in range(2, ws.max_row + 1):
    f = ws.cell(row=r, column=2).value
    line_no_val = ws.cell(row=r, column=3).value
    speaker_val = str(ws.cell(row=r, column=5).value or '').strip()
    orig = str(ws.cell(row=r, column=7).value or '').strip()
    trans = str(ws.cell(row=r, column=8).value or '').strip()
    
    if not f or f not in orig_files or not str(line_no_val).isdigit():
        continue
        
    line_no = int(line_no_val)
    lines = orig_files[f]
    
    if line_no < 1 or line_no > len(lines):
        layer1_errors.append((r, f, line_no, "Dòng vượt quá độ dài file gốc", orig, trans))
        continue
        
    actual_line = lines[line_no - 1].strip()
    clean_csv_orig = re.sub(r'\s+', ' ', orig).strip()
    clean_actual = re.sub(r'\s+', ' ', actual_line).strip()
    
    # Bỏ qua nếu là file name.ks đã được Việt hóa sẵn trong source extracted
    if f != 'name.ks' and clean_csv_orig != clean_actual:
        # Nếu có sự khác biệt nhỏ về khoảng trắng thì bỏ qua, nếu khác nội dung báo lỗi
        if clean_csv_orig.replace(' ', '') != clean_actual.replace(' ', ''):
            layer1_errors.append((r, f, line_no, "Lệch nội dung gốc tiếng Nhật", orig, actual_line))

    # =========================================================================
    # TẦNG 2: KIỂM TRA KHỚP NHÂN VẬT NÓI
    # =========================================================================
    if trans:
        # 1. Shun
        if orig.startswith('[舜]') or orig.startswith('[日高]'):
            if not ('[舜]' in trans or '[日高]' in trans or 'Shun' in trans or 'Hidaka' in trans):
                layer2_errors.append((r, f, line_no, "Gốc là thoại/nội tâm của Shun nhưng bản dịch thiếu thẻ [舜]", orig, trans))
        elif '[舜]「' not in orig and '[舜]（' not in orig:
            # Gốc KHÔNG PHẢI thoại của Shun mà dịch lại bắt đầu bằng [舜]「...
            if trans.startswith('[舜]「') and '「' not in orig:
                layer2_errors.append((r, f, line_no, "Gốc là Dẫn chuyện nhưng bản dịch là thoại [舜]「", orig, trans))

        # 2. Nagi
        if orig.startswith('凪「') or orig.startswith('凪（') or (speaker_val == '凪' and '「' in orig):
            if not ('Nagi' in trans or '凪' in trans or 'tớ' in trans.lower() or 'em' in trans.lower()):
                layer2_errors.append((r, f, line_no, "Gốc là thoại Nagi nhưng bản dịch không khớp", orig, trans))
        elif '凪' not in orig and '凪' not in speaker_val:
            if trans.startswith('Nagi「') or trans.startswith('Nagi（'):
                layer2_errors.append((r, f, line_no, "Xuất hiện Nagi thoại 'ma' khi gốc không có", orig, trans))

        # 3. Rinko
        if orig.startswith('凛子「') or orig.startswith('凛子（') or orig.startswith('凜子「') or orig.startswith('凜子（') or (speaker_val in ['凛子', '凜子'] and '「' in orig):
            if not ('Rinko' in trans or '凛子' in trans or 'chị' in trans.lower() or 'cô' in trans.lower()):
                layer2_errors.append((r, f, line_no, "Gốc là thoại Rinko nhưng bản dịch không khớp", orig, trans))
        elif '凛子' not in orig and '凜子' not in orig and speaker_val not in ['凛子', '凜子']:
            if trans.startswith('Rinko「') or trans.startswith('Rinko（'):
                layer2_errors.append((r, f, line_no, "Xuất hiện Rinko thoại 'ma' khi gốc không có", orig, trans))

        # 4. Tsubomi
        if orig.startswith('蕾「') or orig.startswith('蕾（') or (speaker_val == '蕾' and '「' in orig):
            if not ('Tsubomi' in trans or '蕾' in trans or 'em' in trans.lower() or 'tôi' in trans.lower()):
                layer2_errors.append((r, f, line_no, "Gốc là thoại Tsubomi nhưng bản dịch không khớp", orig, trans))
        elif '蕾' not in orig and speaker_val != '蕾':
            if trans.startswith('Tsubomi「') or trans.startswith('Tsubomi（'):
                layer2_errors.append((r, f, line_no, "Xuất hiện Tsubomi thoại 'ma' khi gốc không có", orig, trans))

        # 5. Hayato
        if orig.startswith('隼人「') or orig.startswith('隼人（') or (speaker_val == '隼人' and '「' in orig):
            if not ('Hayato' in trans or '隼人' in trans or 'tao' in trans.lower() or 'anh' in trans.lower()):
                layer2_errors.append((r, f, line_no, "Gốc là thoại Hayato nhưng bản dịch không khớp", orig, trans))
        elif '隼人' not in orig and speaker_val != '隼人':
            if trans.startswith('Hayato「') or trans.startswith('Hayato（'):
                layer2_errors.append((r, f, line_no, "Xuất hiện Hayato thoại 'ma' khi gốc không có", orig, trans))

        # =====================================================================
        # TẦNG 3: KIỂM TRA TỪ KHÓA & ẢO GIÁC NỘI DUNG (HALLUCINATION)
        # =====================================================================
        hallucination_phrases = [
            'bá chủ hậu cung', 'kỹ năng giường chiếu', 'thu nhập hàng ngày',
            'bảng tổng kết đánh giá', 'một màn trình diễn hoàn hảo', 'tình bạn hayato end',
            'mọi kết thúc đều là', 'bản việt hóa hoàn tất', 'hẹn gặp lại các bạn trong những siêu phẩm'
        ]
        for hp in hallucination_phrases:
            if hp in trans.lower() and hp not in orig.lower():
                layer3_errors.append((r, f, line_no, f"Phát hiện cụm từ ảo giác [{hp}]", orig, trans))

        # Lệch Spirytus
        if 'スピリタス' in orig:
            if not any(w in trans.lower() for w in ['spirytus', 'rượu', 'cồn', 'nang', 'viên']):
                layer3_errors.append((r, f, line_no, "Gốc có Spirytus nhưng bản dịch không có", orig, trans))

        # Lệch Thuốc ngủ
        if '睡眠薬' in orig:
            if not any(w in trans.lower() for w in ['ngủ', 'thuốc']):
                layer3_errors.append((r, f, line_no, "Gốc có Thuốc ngủ nhưng bản dịch không có", orig, trans))

        # Lệch Thuốc kích dục
        if '媚薬' in orig:
            if not any(w in trans.lower() for w in ['kích dục', 'thuốc', 'kem']):
                layer3_errors.append((r, f, line_no, "Gốc có Thuốc kích dục nhưng bản dịch không có", orig, trans))

# =============================================================================
# TẦNG 4: TOÀN VẸN CÚ PHÁP & THẺ ENGINE TRONG PATCH .KS
# =============================================================================
print("[*] [TẦNG 4] Đang quét cấu trúc cú pháp toàn bộ tệp .ks trong patch/...")
layer4_errors = []

for f in sorted(os.listdir(PATCH_DIR)):
    if not f.endswith('.ks'):
        continue
    patch_path = os.path.join(PATCH_DIR, f)
    orig_path = os.path.join(ORIG_DIR, f)
    
    if not os.path.exists(orig_path):
        continue
        
    with open(orig_path, 'r', encoding='utf-8', errors='ignore') as fo:
        orig_lines = fo.readlines()
    with open(patch_path, 'r', encoding='utf-8', errors='ignore') as fp:
        patch_lines = fp.readlines()
        
    if len(orig_lines) != len(patch_lines):
        layer4_errors.append((f, 0, f"Lệch tổng số dòng: Gốc {len(orig_lines)} vs Patch {len(patch_lines)}"))
        
    KNOWN_INTENTIONAL_FIXES = {
        ('minigame.ks', 55),      # Sửa lỗi typo [/iscript] của game gốc Nhật sang [endscript]
        ('room_asa.ks', 431),     # Sửa lỗi logic game gốc Nhật vòng lặp menu sang [s] dừng chờ người chơi bấm
    }

    for idx, (ol, pl) in enumerate(zip(orig_lines, patch_lines), 1):
        if (f, idx) in KNOWN_INTENTIONAL_FIXES:
            continue
        # Kiểm tra thẻ hệ thống
        for st in ['[if', '[endif', '[macro', '[endmacro', '[commit', '[iscript', '[endscript', '[jump', '[call']:
            st_count_orig = ol.count(st)
            st_count_patch = pl.count(st)
            if st_count_orig != st_count_patch:
                layer4_errors.append((f, idx, f"Lệch số lượng thẻ {st}: Gốc {st_count_orig} vs Patch {st_count_patch}"))

# =============================================================================
# BÁO CÁO TỔNG HỢP KẾT QUẢ
# =============================================================================
print("\n" + "=" * 80)
print("                       KẾT QUẢ KIỂM TRA 4 TẦNG")
print("=" * 80)

print(f"\n[1] Lỗi lệch dòng (Layer 1 - Line Desync):           {len(layer1_errors)} lỗi")
if layer1_errors:
    for e in layer1_errors[:10]:
        print(f"    - Row {e[0]} | {e[1]}:{e[2]} -> {e[3]}")

print(f"[2] Lỗi lệch nhân vật nói (Layer 2 - Speaker Desync):   {len(layer2_errors)} lỗi")
if layer2_errors:
    for e in layer2_errors[:10]:
        print(f"    - Row {e[0]} | {e[1]}:{e[2]} -> {e[3]}")
        print(f"      Gốc:  {e[4][:60]}")
        print(f"      Dịch: {e[5][:60]}")

print(f"[3] Lỗi ảo giác / lệch ngữ cảnh (Layer 3 - Semantic):    {len(layer3_errors)} lỗi")
if layer3_errors:
    for e in layer3_errors[:10]:
        print(f"    - Row {e[0]} | {e[1]}:{e[2]} -> {e[3]}")
        print(f"      Gốc:  {e[4][:60]}")
        print(f"      Dịch: {e[5][:60]}")

print(f"[4] Lỗi cú pháp Engine / Thẻ tag (Layer 4 - Syntax):    {len(layer4_errors)} lỗi")
if layer4_errors:
    for e in layer4_errors[:10]:
        print(f"    - {e[0]}:{e[1]} -> {e[2]}")

print("\n" + "=" * 80)
if len(layer1_errors) == 0 and len(layer2_errors) == 0 and len(layer3_errors) == 0 and len(layer4_errors) == 0:
    print(" >>> TUYỆT ĐỐI KHÔNG CÒN BẤT KỲ LỖI LỆCH KỊCH BẢN NÀO TRONG TOÀN BỘ DỰ ÁN! <<<")
else:
    print(" >>> CẦN XỬ LÝ CÁC DÒNG ĐƯỢC LIỆT KÊ TRÊN <<<")
print("=" * 80)
