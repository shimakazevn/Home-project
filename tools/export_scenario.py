#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script xuất dữ liệu kịch bản sang CSV và Excel (XLSX) chuẩn Google Sheets
Dự án: HOME - RJ01556529
Tác giả: VN Patch Pipeline
"""

import os
import sys
import re
import csv
import json
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SCENARIO_DIR = os.path.join(PROJECT_ROOT, 'extracted_scripts', 'data', 'scenario')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'translation')

os.makedirs(OUTPUT_DIR, exist_ok=True)
CSV_OUT = os.path.join(OUTPUT_DIR, 'text_export.csv')
XLSX_OUT = os.path.join(OUTPUT_DIR, 'text_export.xlsx')
UI_CSV_OUT = os.path.join(OUTPUT_DIR, 'ui_export.csv')   # File UI riêng
META_OUT = os.path.join(OUTPUT_DIR, 'export_metadata.json')

# Entry types đưa vào file chính (cần dịch bằng AI)
MAIN_TYPES = {'dialogue', 'narration'}
# Entry types đưa ra file UI riêng (tự dịch được)
UI_TYPES = {'glink_choice', 'ptext_ui'}


def clean_speaker_name(raw_name, line_text):
    """Chuẩn hóa tên người nói để hiển thị rõ ràng trên bảng tính."""
    raw = raw_name.strip()
    if not raw and ('（' in line_text or '(' in line_text) and '「' not in line_text:
        return 'Nam chính (Nội tâm)'
    if not raw:
        return 'Dẫn chuyện'
    
    name_map = {
        '#凪': '月城 凪 (Nagi)',
        '凪': '月城 凪 (Nagi)',
        '#凛子': '成瀬 凛子 (Rinko)',
        '凛子': '成瀬 凛子 (Rinko)',
        '#蕾': '成瀬 蕾 (Tsubomi)',
        '蕾': '成瀬 蕾 (Tsubomi)',
        '#隼人': '成瀬 隼人 (Hayato)',
        '隼人': '成瀬 隼人 (Hayato)',
        '#舜': '日高 舜 (MC)',
        '[舜]': '日高 舜 (MC)',
        '舜': '日高 舜 (MC)',
        '[日高]': '日高 舜 (MC)',
        '日高': '日高 舜 (MC)',
        '？？？': 'Người lạ (???)',
        '#アメリア': 'Amelia (Tsubomi)',
        'アメリア': 'Amelia (Tsubomi)',
        '#田中': 'Bác Tanaka (Bệnh nhân)',
        '田中': 'Bác Tanaka (Bệnh nhân)',
        '#客': 'Khách hàng',
        '客': 'Khách hàng',
        '#店員': 'Nhân viên cửa hàng',
        '店員': 'Nhân viên cửa hàng',
        '#男': 'Người đàn ông',
        '#警察': 'Cảnh sát',
    }
    
    return name_map.get(raw, raw.lstrip('#').strip('[]'))


def get_category_order(filename):
    """Xác định thứ tự danh mục theo cốt truyện để sắp xếp file logic."""
    fl = filename.lower()
    if fl.startswith('ev_op') or fl in ['name.ks', 'first.ks', 'scene1.ks']:
        return (1, 'Story_Opening')
    elif fl.startswith('room_') or fl.startswith('ev_nagi') or fl.startswith('ev_rinko') or fl.startswith('ev_tubomi') or fl in ['limit.ks', 'ev_honmei.ks']:
        return (2, 'Story_Main')
    elif fl.startswith('komyu_') or fl.startswith('komyuhyouka_'):
        return (3, 'Story_Commu')
    elif fl.startswith('sinnyu_') or fl.startswith('ev_trap_'):
        return (4, 'Story_Infiltration')
    elif fl.startswith('job_') or fl.startswith('mesi_'):
        return (5, 'Story_Work')
    elif fl.startswith('sansaku_'):
        return (6, 'Story_Exploration')
    elif fl.startswith('ev_date') or fl.startswith('ev_deto'):
        return (7, 'Story_Date')
    elif fl.startswith('ev_kokuhaku') or fl in ['ev_hukusyu_hayato.ks', 'ev_omoidenokakera.ks']:
        return (8, 'Story_Confession')
    elif fl.startswith('h_'):
        return (9, '18+_HScene')
    elif 'badend' in fl:
        return (11, 'Ending_BadEnd')
    elif 'end' in fl or 'ed' in fl:
        return (10, 'Ending_GoodEnd')
    else:
        return (12, 'UI_System')


def get_context_notes(speaker_clean, scene_type, line_text):
    """Tạo ghi chú ngữ cảnh và gợi ý xưng hô chuẩn."""
    notes = []
    if 'Nagi' in speaker_clean:
        if '18+' in scene_type:
            notes.append('Nagi nói trong cảnh 18+: giọng dâm đê mê, xưng em - anh Shun')
        else:
            notes.append('Nagi nói chuyện: dịu dàng, xưng tớ - Shun-chan (hoặc em - anh)')
    elif 'Rinko' in speaker_clean:
        if '18+' in scene_type:
            notes.append('Mẹ Rinko cảnh 18+: dâm đãng, tội lỗi, xưng em - anh (hoặc chị - em)')
        else:
            notes.append('Mẹ Rinko: lịch thiệp, đằm thắm, xưng cô - cháu (hoặc tôi/chị - Shun-kun)')
    elif 'Tsubomi' in speaker_clean or 'Amelia' in speaker_clean:
        if '18+' in scene_type:
            notes.append('Tsubomi cảnh 18+: ngây thơ, rên rỉ, xưng em - anh')
        elif 'Amelia' in speaker_clean:
            notes.append('Tsubomi đóng vai Maid Amelia: xưng em - Quý khách/Chủ nhân')
        else:
            notes.append('Tsubomi: tsundere, đanh đá, xưng em - anh (hoặc tôi - anh)')
    elif 'Hayato' in speaker_clean:
        notes.append('Hayato: tự phụ, trịch thượng, xưng tôi - cậu / anh - chú mày')
    elif 'Nam chính' in speaker_clean:
        notes.append('Shun (MC) độc thoại nội tâm: xưng tôi hoặc mình, toan tính mưu mô')

    if '[r]' in line_text:
        notes.append('Có tag [r] xuống dòng')
    if '[p]' in line_text:
        notes.append('Có tag [p] chuyển trang')
    if 'exp=' in line_text or 'emb' in line_text:
        notes.append('Bảo toàn nguyên vẹn mã biến [emb exp="..."]')
        
    return ' | '.join(notes)


def extract_all():
    """Trích xuất toàn bộ dữ liệu từ 267 tệp .ks với logic lọc sạch và phân loại."""
    raw_files = [f for f in os.listdir(SCENARIO_DIR) if f.endswith('.ks')]
    
    # Sắp xếp file theo thứ tự cốt truyện logic
    sorted_files = sorted(raw_files, key=lambda x: (get_category_order(x)[0], x))
    print(f"Bắt đầu quét {len(sorted_files)} tệp .ks trong {SCENARIO_DIR}...")
    
    entries = []
    metadata = {}
    row_id = 0
    
    for f in sorted_files:
        path = os.path.join(SCENARIO_DIR, f)
        lines = open(path, 'r', encoding='utf-8', errors='replace').readlines()
        
        in_tb_text = False
        in_iscript = False
        in_tyrano_code = False
        in_html = False
        in_macro = False
        current_speaker_raw = ""
        current_speaker_clean = ""
        cat_num, scene_type = get_category_order(f)
        file_entries = []
        
        for line_idx, raw_line in enumerate(lines):
            line_num = line_idx + 1
            line = raw_line.strip()
            
            # 1. Bỏ qua dòng trống, comment (; hoặc // hoặc /*)
            if not line or line.startswith(';') or line.startswith('//') or line.startswith('/*'):
                continue
                
            # 2. Theo dõi và chuyển đổi trạng thái các khối code / engine
            if '[iscript]' in line:
                in_iscript = True
                continue
            if '[endscript]' in line:
                in_iscript = False
                continue
            if '[tb_start_tyrano_code' in line:
                in_tyrano_code = True
                continue
            if '[_tb_end_tyrano_code' in line:
                in_tyrano_code = False
                continue
            if '[html]' in line:
                in_html = True
                continue
            if '[endhtml]' in line:
                in_html = False
                continue
            if '[macro' in line:
                in_macro = True
                continue
            if '[endmacro]' in line:
                in_macro = False
                continue

            # NẾU ĐANG NẰM TRONG CODE BLOCK -> BỎ QUA 100% ĐỂ TRÁNH HỎNG CODE GAME
            if in_iscript or in_tyrano_code or in_html or in_macro:
                continue

            # 3. Bỏ qua các ký tự đóng mở tag thừa / code syntax
            if line in [']', '[', '}}', '{', '}', '};', ');', '})', '});']:
                continue
                
            # 4. Speaker tag #Name (cập nhật tên người nói, không xuất thành dòng dịch riêng)
            if line.startswith('#'):
                spk = line[1:].strip()
                if spk:
                    current_speaker_raw = line
                    current_speaker_clean = clean_speaker_name(spk, "")
                else:
                    current_speaker_raw = ""
                    current_speaker_clean = ""
                continue
                
            # 5. tb_start_text / _tb_end_text
            if '[tb_start_text' in line:
                in_tb_text = True
                continue
            if '[_tb_end_text' in line:
                in_tb_text = False
                continue
                
            # 6. Glink choices
            if 'glink' in line and 'text=' in line:
                match = re.search(r'text=[\"\'](.*?)[\"\']', line)
                if match:
                    choice_text = match.group(1)
                    if choice_text.strip() and not choice_text.strip() in [']', '[']:
                        row_id += 1
                        entry = {
                            'row_id': row_id,
                            'file': f,
                            'line_number': line_num,
                            'entry_type': 'glink_choice',
                            'speaker_clean': 'Lựa chọn',
                            'speaker_raw': '',
                            'original_jp': choice_text,
                            'vietnamese': '',
                            'tag_codes': '',
                            'scene_type': 'UI_Choice',
                            'notes': 'Lựa chọn tương tác (glink) - Dịch ngắn gọn, chuẩn nghĩa',
                            'line_template': line
                        }
                        entries.append(entry)
                        file_entries.append(entry)
                continue

            # 7. ptext (static display text)
            if 'ptext' in line and 'text=' in line:
                match = re.search(r'text=[\"\'](.*?)[\"\']', line)
                if match:
                    ptext_val = match.group(1)
                    if not ptext_val.startswith('&') and len(ptext_val.strip()) > 0 and not ptext_val.strip() in [']', '[', '％']:
                        row_id += 1
                        entry = {
                            'row_id': row_id,
                            'file': f,
                            'line_number': line_num,
                            'entry_type': 'ptext_ui',
                            'speaker_clean': 'Hệ thống',
                            'speaker_raw': '',
                            'original_jp': ptext_val,
                            'vietnamese': '',
                            'tag_codes': '',
                            'scene_type': 'UI_System',
                            'notes': 'Văn bản giao diện tĩnh (ptext)',
                            'line_template': line
                        }
                        entries.append(entry)
                        file_entries.append(entry)
                continue

            # 8. Text inside [tb_start_text] block
            if in_tb_text:
                stripped_tags = re.sub(r'\[.*?\]', '', line).strip()
                if not stripped_tags and ('[' in line and ']' in line):
                    continue
                    
                spk_c = current_speaker_clean
                spk_r = current_speaker_raw
                
                spk_bracket = re.match(r'^(?:\[(.*?)\]|([^\s\[「」]+))「', line)
                if spk_bracket:
                    name_found = spk_bracket.group(1) or spk_bracket.group(2)
                    if name_found:
                        spk_c = clean_speaker_name(name_found, line)
                        spk_r = f"[{name_found}]" if spk_bracket.group(1) else name_found

                if not spk_c:
                    spk_c = clean_speaker_name("", line)

                tags = re.findall(r'\[.*?\]', line)
                tag_str = ', '.join(tags) if tags else ''
                notes = get_context_notes(spk_c, scene_type, line)
                
                row_id += 1
                entry = {
                    'row_id': row_id,
                    'file': f,
                    'line_number': line_num,
                    'entry_type': 'dialogue' if ('「' in line or (spk_r and spk_r != '#')) else 'narration',
                    'speaker_clean': spk_c,
                    'speaker_raw': spk_r,
                    'original_jp': line,
                    'vietnamese': '',
                    'tag_codes': tag_str,
                    'scene_type': scene_type,
                    'notes': notes,
                    'line_template': line
                }
                entries.append(entry)
                file_entries.append(entry)
                continue

            # 9. Standalone text lines outside tb_start_text (chỉ lấy thoại / dẫn truyện thuần)
            if not line.startswith('[') and not line.startswith('*') and not line.startswith('#'):
                # Kiểm tra ký tự tiếng Nhật hoặc ngoặc thoại
                has_jp = bool(re.search(r'[\u3040-\u309f\u30a0-\u30ff\u4e00-\u9faf]', line))
                has_dialogue_bracket = '「' in line and '」' in line
                
                # Bộ lọc chặn mã code JavaScript / jQuery / CSS lọt vào
                is_js_code = any(kw in line for kw in [
                    'var ', 'let ', 'const ', 'function(', 'function (', 'return ',
                    '$("', "$('", '.css(', '.on(', '.append(', '.remove(', 'TYRANO.', 'Math.',
                    'position:', 'right:', 'bottom:', 'width:', 'height:', 'lineHeight:',
                    'background:', 'border:', 'fontSize:', 'zIndex:', 'cursor:'
                ])

                if (has_jp or has_dialogue_bracket) and not is_js_code:
                    spk_c = current_speaker_clean
                    spk_r = current_speaker_raw
                    
                    spk_bracket = re.match(r'^(?:\[(.*?)\]|([^\s\[「」]+))「', line)
                    if spk_bracket:
                        name_found = spk_bracket.group(1) or spk_bracket.group(2)
                        if name_found:
                            spk_c = clean_speaker_name(name_found, line)
                            spk_r = f"[{name_found}]" if spk_bracket.group(1) else name_found

                    if not spk_c:
                        spk_c = clean_speaker_name("", line)

                    tags = re.findall(r'\[.*?\]', line)
                    tag_str = ', '.join(tags) if tags else ''
                    notes = get_context_notes(spk_c, scene_type, line)
                    
                    row_id += 1
                    entry = {
                        'row_id': row_id,
                        'file': f,
                        'line_number': line_num,
                        'entry_type': 'dialogue' if ('「' in line or (spk_r and spk_r != '#')) else 'narration',
                        'speaker_clean': spk_c,
                        'speaker_raw': spk_r,
                        'original_jp': line,
                        'vietnamese': '',
                        'tag_codes': tag_str,
                        'scene_type': scene_type,
                        'notes': notes,
                        'line_template': line
                    }
                    entries.append(entry)
                    file_entries.append(entry)

        metadata[f] = {
            'total_lines': len(lines),
            'entry_count': len(file_entries)
        }

    print(f"Tổng cộng đã trích xuất: {len(entries)} dòng cần dịch từ {len(sorted_files)} tệp.")
    return entries, metadata


def write_csv(entries):
    """Ghi ra CSV chuẩn UTF-8 with BOM, RFC 4180 quoting, chống lệch dòng.
    Chỉ xuất entry_type thuộc MAIN_TYPES (dialogue + narration).
    """
    headers = [
        'row_id', 'file', 'line_number', 'entry_type',
        'speaker_clean', 'speaker_raw', 'original_jp',
        'vietnamese', 'tag_codes', 'scene_type', 'notes'
    ]
    main_entries = [e for e in entries if e['entry_type'] in MAIN_TYPES]
    
    with open(CSV_OUT, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(
            f, fieldnames=headers,
            quoting=csv.QUOTE_ALL,
            extrasaction='ignore'
        )
        writer.writeheader()
        for e in main_entries:
            row_copy = dict(e)
            for k in ['original_jp', 'vietnamese', 'notes', 'speaker_clean']:
                val = row_copy.get(k, '')
                if val and str(val)[0] in ('=', '+', '-', '@'):
                    row_copy[k] = "'" + str(val)
            writer.writerow(row_copy)

    print(f"[OK] CSV chính: {CSV_OUT} ({os.path.getsize(CSV_OUT):,} bytes) -- {len(main_entries)} dòng")


def write_ui_csv(entries):
    """Xuất riêng file UI (glink_choice + ptext_ui) ra ui_export.csv.
    File này có cấu trúc đơn giản hơn, dễ dịch thủ công.
    """
    ui_entries = [e for e in entries if e['entry_type'] in UI_TYPES]
    if not ui_entries:
        print("[INFO] Không có entry UI nào để xuất.")
        return

    headers = ['row_id', 'file', 'line_number', 'entry_type', 'original_jp', 'vietnamese', 'notes']
    with open(UI_CSV_OUT, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_ALL, extrasaction='ignore')
        writer.writeheader()
        for e in ui_entries:
            row_copy = dict(e)
            val = row_copy.get('original_jp', '')
            if val and str(val)[0] in ('=', '+', '-', '@'):
                row_copy['original_jp'] = "'" + str(val)
            writer.writerow(row_copy)

    print(f"[OK] CSV UI riêng: {UI_CSV_OUT} ({os.path.getsize(UI_CSV_OUT):,} bytes) -- {len(ui_entries)} dòng")


def write_xlsx(entries):
    """Ghi ra Excel (.xlsx) với định dạng đẹp, đóng băng hàng tiêu đề, tự động căn chỉnh độ rộng cột."""
    if not HAS_OPENPYXL:
        print("[WARN] openpyxl chưa cài đặt, bỏ qua xuất XLSX.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HOME_Translation"

    # Styling definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_thin = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    headers = [
        'row_id', 'file', 'line_number', 'entry_type',
        'speaker_clean', 'speaker_raw', 'original_jp',
        'vietnamese', 'tag_codes', 'scene_type', 'notes'
    ]

    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Chỉ đưa dialogue + narration vào sheet chính
    main_entries = [e for e in entries if e['entry_type'] in MAIN_TYPES]

    for r_idx, e in enumerate(main_entries, start=2):
        row_data = [
            e['row_id'], e['file'], e['line_number'], e['entry_type'],
            e['speaker_clean'], e['speaker_raw'], e['original_jp'],
            e['vietnamese'], e['tag_codes'], e['scene_type'], e['notes']
        ]
        ws.append(row_data)

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=r_idx, column=col_idx)
            c.font = data_font
            c.border = border_thin
            if headers[col_idx - 1] in ['original_jp', 'vietnamese', 'notes']:
                c.alignment = Alignment(vertical="top", wrap_text=True)
            elif headers[col_idx - 1] in ['row_id', 'line_number']:
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(horizontal="left", vertical="top")

    ws.freeze_panes = "A2"

    col_widths = {
        'A': 10,  # row_id
        'B': 25,  # file
        'C': 12,  # line_number
        'D': 15,  # entry_type
        'E': 22,  # speaker_clean
        'F': 15,  # speaker_raw
        'G': 55,  # original_jp
        'H': 55,  # vietnamese
        'I': 22,  # tag_codes
        'J': 18,  # scene_type
        'K': 40   # notes
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(XLSX_OUT)
    main_count = sum(1 for e in entries if e['entry_type'] in MAIN_TYPES)
    print(f"[OK] Excel chính: {XLSX_OUT} ({os.path.getsize(XLSX_OUT):,} bytes) -- {main_count} dòng")


def save_metadata(metadata):
    """Lưu metadata cấu trúc kịch bản để hỗ trợ reimport tuyệt đối chính xác."""
    with open(META_OUT, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)
    print(f"[OK] Đã lưu metadata: {META_OUT}")


if __name__ == '__main__':
    entries, metadata = extract_all()

    main_count = sum(1 for e in entries if e['entry_type'] in MAIN_TYPES)
    ui_count = sum(1 for e in entries if e['entry_type'] in UI_TYPES)
    print(f"  >> Script chính (dialogue + narration): {main_count} dòng")
    print(f"  >> UI riêng (glink_choice + ptext_ui): {ui_count} dòng")
    print()

    write_csv(entries)
    write_ui_csv(entries)
    write_xlsx(entries)
    save_metadata(metadata)
    print("\n>>> HOÀN TẤT XUẤT TOÀN BỘ DỮ LIỆU SANG CSV VÀ EXCEL!")
    print(f"    text_export.xlsx  = {main_count} dòng cần dịch AI")
    print(f"    ui_export.csv     = {ui_count} dòng UI tự dịch")
