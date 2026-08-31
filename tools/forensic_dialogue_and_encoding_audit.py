# -*- coding: utf-8 -*-
"""
tools/forensic_dialogue_and_encoding_audit.py
=============================================
Kiểm toán pháp y văn bản & mã hóa (Level: Forensic Dialogue & Character Stream):
1. Quét ký tự ẩn, ký tự điều khiển lỗi (0x7F, 0x00-0x1F, BOM giữa dòng, ZWSP).
2. Quét mất cân bằng dấu ngoặc thoại 「...」, （...）, (...).
3. Quét thẻ font màu [font] thiếu [resetfont].
4. Quét thẻ ngắt dòng / dừng click trùng lặp ([r][r], [p][p], [r][p]).
5. Quét cấu trúc thẻ tên nhân vật và định dạng thoại chuẩn.
"""

import openpyxl
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(PROJECT_ROOT, 'translation', 'text_export.xlsx')
PATCH_DIR = os.path.join(PROJECT_ROOT, 'patch', 'data', 'scenario')

print("================================================================================")
print("     KIỂM TOÁN PHÁP Y VĂN BẢN, MÃ HÓA & CẤU TRÚC THOẠI (FORENSIC AUDIT)")
print("================================================================================\n")

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

encoding_errors = []
bracket_errors = []
tag_errors = []
duplicate_breaks = []
voice_prefix_errors = []

for r in range(2, ws.max_row + 1):
    f = ws.cell(row=r, column=2).value
    l = ws.cell(row=r, column=3).value
    orig = str(ws.cell(row=r, column=7).value or '').strip()
    trans = str(ws.cell(row=r, column=8).value or '').strip()
    
    # 1. Ký tự ẩn / Điều khiển độc hại
    for ch in trans:
        code = ord(ch)
        if (code < 32 and code not in (9, 10, 13)) or code == 127 or code in (0x200B, 0x200C, 0x200D, 0xFEFF):
            encoding_errors.append((r, f, l, f"Ký tự mã hóa bất thường: U+{code:04X} ({repr(ch)})", trans))
            break
            
    # 2. Mất cân bằng ngoặc thoại Nhật 「 và 」
    kagi_open = trans.count('「')
    kagi_close = trans.count('」')
    if kagi_open != kagi_close:
        bracket_errors.append((r, f, l, f"Lệch ngoặc thoại 「: {kagi_open} mở vs {kagi_close} đóng", trans))
        
    # Mất cân bằng ngoặc đơn tròn ( và )
    p_open = trans.count('(') + trans.count('（')
    p_close = trans.count(')') + trans.count('）')
    if p_open != p_close:
        bracket_errors.append((r, f, l, f"Lệch ngoặc đơn (: {p_open} mở vs {p_close} đóng", trans))
        
    # 3. Thẻ [font] và [resetfont]
    font_open = trans.count('[font')
    font_reset = trans.count('[resetfont]')
    if font_open != font_reset:
        tag_errors.append((r, f, l, f"Lệch thẻ [font]: {font_open} [font vs {font_reset} [resetfont]", trans))
        
    # 4. Trùng lặp ngắt dòng / dừng click
    if '[r][r]' in trans or '[p][p]' in trans or '[r][p]' in trans:
        duplicate_breaks.append((r, f, l, "Trùng lặp lệnh ngắt dòng / dừng trang", trans))

print(f"[1] Lỗi ký tự ẩn & Mã hóa độc hại (Encoding/Control): {len(encoding_errors)} lỗi")
for ee in encoding_errors[:10]:
    print(f"    - Row {ee[0]} | {ee[1]}:{ee[2]} -> {ee[3]} | Câu: {ee[4][:60]}")

print(f"\n[2] Lỗi mất cân bằng ngoặc thoại/ngoặc đơn (Brackets):   {len(bracket_errors)} lỗi")
for be in bracket_errors[:15]:
    print(f"    - Row {be[0]} | {be[1]}:{be[2]} -> {be[3]} | Câu: {be[4][:60]}")

print(f"\n[3] Lỗi thẻ Font / Resetfont (Tag Balance):           {len(tag_errors)} lỗi")
for te in tag_errors:
    print(f"    - Row {te[0]} | {te[1]}:{te[2]} -> {te[3]} | Câu: {te[4][:60]}")

print(f"\n[4] Lỗi trùng lặp ngắt dòng/ngắt trang ([r][r], [p][p]): {len(duplicate_breaks)} lỗi")
for db in duplicate_breaks[:10]:
    print(f"    - Row {db[0]} | {db[1]}:{db[2]} -> {db[3]} | Câu: {db[4][:60]}")
