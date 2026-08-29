#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CHROME AUTO-PROMPT BOT CHO GOOGLE WORKSPACE / GEMINI
Tự động gửi batch kịch bản từ text_export.xlsx vào Gemini trên Chrome và lưu kết quả
Dự án: HOME (ROOM) - RJ01556529
"""

import os
import sys
import re
import json
import time
import urllib.request
import urllib.error
import subprocess
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[LỖI] Cần cài đặt openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    import websocket
except ImportError:
    try:
        import websockets
    except ImportError:
        print("[LỖI] Cần cài đặt websocket-client: pip install websocket-client")
        sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TRANSLATION_DIR = PROJECT_ROOT / 'translation'
XLSX_PATH = TRANSLATION_DIR / 'text_export.xlsx'
PROMPT_FILE = PROJECT_ROOT / 'GEMINI_WORKSPACE_PROMPT.md'
CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"

BATCH_SIZE = 25  # Số dòng mỗi lần prompt


def get_chrome_executable():
    """Tìm file thực thi của Google Chrome trên Windows."""
    candidates = [
        os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return "chrome.exe"


def is_chrome_running():
    """Kiểm tra xem Chrome đã bật cổng Remote Debugging chưa."""
    try:
        req = urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
        return req.status == 200
    except Exception:
        return False


def start_chrome_debug():
    """Khởi động Chrome với cổng Remote Debugging."""
    chrome_exe = get_chrome_executable()
    user_data = os.path.expandvars(r"%TEMP%\Chrome_Gemini_Bot_Profile")
    print(f"Đang khởi động Chrome Debugging Mode: {chrome_exe}")
    print(f"Thư mục hồ sơ tạm: {user_data}")
    
    cmd = [
        chrome_exe,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={user_data}",
        "https://gemini.google.com/app"
    ]
    subprocess.Popen(cmd)
    
    print("Đang chờ Chrome sẵn sàng...")
    for _ in range(15):
        time.sleep(1)
        if is_chrome_running():
            print("[OK] Chrome Debugging đã kết nối thành công!")
            return True
            
    print("[LỖI] Không thể kết nối tới Chrome Debugging!")
    return False


def get_gemini_tab():
    """Tìm tab Gemini đang mở hoặc tạo tab mới."""
    try:
        req = urllib.request.urlopen(f"{CDP_URL}/json", timeout=5)
        tabs = json.loads(req.read().decode('utf-8'))
        
        # Tìm tab Gemini
        for t in tabs:
            if t.get('type') == 'page' and ('gemini.google.com' in t.get('url', '') or 'google' in t.get('url', '')):
                return t
                
        # Nếu chưa có tab Gemini, lấy tab đầu tiên hoặc tạo mới
        if tabs:
            return tabs[0]
    except Exception as e:
        print(f"[LỖI] Không lấy được danh sách tab: {e}")
    return None


class CDPClient:
    """Client WebSocket giao tiếp trực tiếp với Chrome DevTools Protocol."""
    def __init__(self, ws_url):
        import websocket
        self.ws = websocket.create_connection(ws_url, timeout=30)
        self.req_id = 0

    def send_cmd(self, method, params=None):
        self.req_id += 1
        payload = {"id": self.req_id, "method": method, "params": params or {}}
        self.ws.send(json.dumps(payload))
        while True:
            res = json.loads(self.ws.recv())
            if res.get("id") == self.req_id:
                return res.get("result", {})

    def evaluate(self, js_expr):
        """Thực thi mã JavaScript trên trang web."""
        res = self.send_cmd("Runtime.evaluate", {
            "expression": js_expr,
            "returnByValue": True,
            "awaitPromise": True
        })
        return res.get("result", {}).get("value")

    def close(self):
        try:
            self.ws.close()
        except Exception:
            pass


def build_batch_prompt(system_rules, rows):
    """Tạo prompt chuẩn JSON cho batch dòng cần dịch."""
    items = []
    for r in rows:
        items.append({
            "id": r["row_id"],
            "speaker": r["speaker_clean"],
            "text": r["original_jp"],
            "scene": r["scene_type"]
        })
        
    prompt = f"""{system_rules}

---

### YÊU CẦU DỊCH BATCH DỮ LIỆU DƯỚI ĐÂY
Hãy dịch chính xác các câu thoại sau sang tiếng Việt, tuân thủ 100% Ma trận xưng hô, Bảng từ cấm 18+ và BẢO TOÀN NGUYÊN VẸN các thẻ lệnh trong `[...]` (như `[r]`, `[p]`, `[emb ...]`, `[舜]`).

ĐẦU VÀO JSON:
```json
{json.dumps(items, ensure_ascii=False, indent=2)}
```

TRẢ VỀ KẾT QUẢ THEO ĐÚNG ĐỊNH DẠNG JSON MẢNG CÁC ĐỐI TƯỢNG (BẮT BUỘC 1:1, KHÔNG BỎ SÓT BẤT KỲ ID NÀO):
```json
[
  {{"id": 1, "vietnamese": "Nội dung dịch tiếng Việt..."}},
  ...
]
```
Chỉ trả về khối mã JSON duy nhất, không kèm giải thích bên ngoài."""
    return prompt


def run_bot():
    print("=" * 65)
    print("  BOT AUTO-PROMPT GOOGLE WORKSPACE / GEMINI CHO EXCEL TRANSLATION")
    print("=" * 65)

    if not XLSX_PATH.exists():
        print(f"[LỖI] Không tìm thấy file {XLSX_PATH}")
        return

    # 1. Khởi động / Kết nối Chrome
    if not is_chrome_running():
        print("Chrome chưa bật chế độ Debugging. Đang tự động kích hoạt...")
        if not start_chrome_debug():
            return
    else:
        print("[OK] Đã kết nối với Chrome đang mở trên cổng 9222.")

    # 2. Lấy tab Gemini
    tab = get_gemini_tab()
    if not tab or not tab.get('webSocketDebuggerUrl'):
        print("[LỖI] Không tìm thấy tab nào có WebSocket Debugger URL!")
        return

    print(f"Đang kết nối vào tab: {tab.get('title', 'Untitled')} ({tab.get('url')})")
    cdp = CDPClient(tab['webSocketDebuggerUrl'])

    # 3. Đọc System Instruction
    system_rules = ""
    if PROMPT_FILE.exists():
        system_rules = open(PROMPT_FILE, 'r', encoding='utf-8').read()

    # 4. Đọc dữ liệu Excel
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]
    
    # Tìm các hàng chưa dịch
    untranslated_rows = []
    for row_idx in range(2, ws.max_row + 1):
        r_dict = {}
        for col_idx, h in enumerate(headers, start=1):
            r_dict[h] = ws.cell(row=row_idx, column=col_idx).value
        
        vn_val = str(r_dict.get('vietnamese') or '').strip()
        if not vn_val or vn_val.lower() == 'none':
            r_dict['_row_num'] = row_idx
            untranslated_rows.append(r_dict)

    total_untranslated = len(untranslated_rows)
    print(f"\nTổng số dòng cần dịch: {total_untranslated:,} dòng.")
    if total_untranslated == 0:
        print("[THÔNG BÁO] Toàn bộ file đã được dịch 100%!")
        return

    # 5. Xử lý từng batch
    batch_num = 0
    total_batches = (total_untranslated + BATCH_SIZE - 1) // BATCH_SIZE

    for i in range(0, total_untranslated, BATCH_SIZE):
        batch_num += 1
        chunk = untranslated_rows[i:i + BATCH_SIZE]
        print(f"\n--- [BATCH {batch_num}/{total_batches}] Đang xử lý {len(chunk)} dòng ({chunk[0]['row_id']} -> {chunk[-1]['row_id']}) ---")

        prompt_text = build_batch_prompt(system_rules, chunk)

        # Gửi prompt vào khung chat Gemini bằng JavaScript DOM
        js_inject_and_send = f"""
        (function() {{
            const prompt = {json.dumps(prompt_text)};
            
            // Tìm khung nhập liệu của Gemini
            let textarea = document.querySelector('rich-textarea p') || 
                           document.querySelector('div[contenteditable="true"]') ||
                           document.querySelector('textarea');
                           
            if (!textarea) return {{ success: false, error: "Không tìm thấy ô nhập chat" }};
            
            textarea.focus();
            textarea.innerText = prompt;
            textarea.dispatchEvent(new Event('input', {{ bubbles: true }}));
            
            // Tìm nút gửi
            setTimeout(() => {{
                let sendBtn = document.querySelector('button[aria-label*="Send"]') || 
                              document.querySelector('button[aria-label*="Gửi"]') ||
                              document.querySelector('.send-button') ||
                              document.querySelector('button[jsname="Qx7uuf"]');
                              
                if (sendBtn && !sendBtn.disabled) {{
                    sendBtn.click();
                }}
            }}, 500);
            
            return {{ success: true }};
        }})();
        """
        res = cdp.evaluate(js_inject_and_send)
        if not res or not res.get('success'):
            print(f"[CẢNH BÁO] Không thể tự động bấm nút gửi: {res}")
            print("Vui lòng kiểm tra tab Gemini trên trình duyệt.")

        print("Đang chờ Gemini hoàn tất câu trả lời...")
        
        # Chờ Gemini sinh câu trả lời xong
        translated_data = None
        for wait_sec in range(60):
            time.sleep(3)
            # Lấy câu trả lời mới nhất từ DOM
            js_get_last_response = """
            (function() {
                // Kiểm tra xem có nút Stop đang hiện không (đang streaming)
                let isGenerating = !!document.querySelector('button[aria-label*="Stop"]') || 
                                   !!document.querySelector('button[aria-label*="Dừng"]');
                
                // Lấy tất cả khối phản hồi
                let responses = document.querySelectorAll('.model-response-text, message-content, .response-content');
                if (responses.length === 0) return { status: "waiting", text: "" };
                
                let lastResponse = responses[responses.length - 1].innerText;
                return {
                    status: isGenerating ? "generating" : "done",
                    text: lastResponse
                };
            })();
            """
            check = cdp.evaluate(js_get_last_response)
            if check and check.get('status') == 'done' and len(check.get('text', '')) > 50:
                resp_text = check['text']
                # Trích xuất JSON từ markdown block
                json_match = re.search(r'```(?:json)?\s*(\[\s*\{.*?\}\s*\])\s*```', resp_text, re.DOTALL)
                if json_match:
                    try:
                        translated_data = json.loads(json_match.group(1))
                        break
                    except Exception:
                        pass
                else:
                    # Thử parse trực tiếp mảng JSON
                    arr_match = re.search(r'(\[\s*\{.*?\}\s*\])', resp_text, re.DOTALL)
                    if arr_match:
                        try:
                            translated_data = json.loads(arr_match.group(1))
                            break
                        except Exception:
                            pass

        if translated_data:
            print(f"[OK] Đã nhận {len(translated_data)} câu dịch từ Gemini!")
            # Ghi vào file Excel
            trans_map = {int(item.get('id', 0)): str(item.get('vietnamese', '')).strip() for item in translated_data if item.get('id')}
            
            applied_count = 0
            for r in chunk:
                rid = int(r['row_id'])
                if rid in trans_map and trans_map[rid]:
                    ws.cell(row=r['_row_num'], column=8, value=trans_map[rid])
                    applied_count += 1

            wb.save(XLSX_PATH)
            print(f"[OK] Đã lưu {applied_count}/{len(chunk)} dòng vào {XLSX_PATH.name} thành công!")
        else:
            print(f"[CẢNH BÁO] Không lấy được JSON phản hồi hợp lệ cho batch {batch_num}. Bỏ qua để xử lý sau.")

        time.sleep(2)

    cdp.close()
    print("\n" + "=" * 65)
    print("  >>> HOÀN TẤT CHẠY BOT DỊCH THUẬT AUTO-PROMPT TRÊN CHROME!")
    print("=" * 65)


if __name__ == '__main__':
    run_bot()
