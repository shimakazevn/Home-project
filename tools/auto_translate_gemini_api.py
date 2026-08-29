#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TOOL DỊCH TỰ ĐỘNG TỐC ĐỘ CAO BẰNG GOOGLE GEMINI API (1.5 / 2.0 FLASH)
Dự án: HOME (ROOM) - RJ01556529
"""

import os
import sys
import re
import json
import time
import urllib.request
import urllib.error
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[LỖI] Cần cài đặt openpyxl: pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TRANSLATION_DIR = PROJECT_ROOT / 'translation'
XLSX_PATH = TRANSLATION_DIR / 'text_export.xlsx'
PROMPT_FILE = PROJECT_ROOT / 'GEMINI_WORKSPACE_PROMPT.md'

# Khóa API Gemini (Lấy từ biến môi trường hoặc nhập trực tiếp)
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
API_MODEL = "gemini-2.0-flash"  # hoặc gemini-1.5-flash
BATCH_SIZE = 30  # Số dòng mỗi batch


def call_gemini_api(api_key, system_instruction, user_content):
    """Gọi trực tiếp Google Gemini REST API qua HTTP POST."""
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{API_MODEL}:generateContent?key={api_key}"
    
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": user_content}]
            }
        ],
        "systemInstruction": {
            "parts": [{"text": system_instruction}]
        },
        "generationConfig": {
            "temperature": 0.2,
            "responseMimeType": "application/json"
        }
    }
    
    data_bytes = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data_bytes, headers={'Content-Type': 'application/json'})
    
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            res_json = json.loads(resp.read().decode('utf-8'))
            candidates = res_json.get('candidates', [])
            if candidates:
                text_out = candidates[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                return json.loads(text_out)
    except Exception as e:
        print(f"  [LỖI API] {e}")
        return None


def run_gemini_translator():
    print("=" * 65)
    print("  DỊCH TỰ ĐỘNG BẰNG GOOGLE GEMINI API (TỐC ĐỘ CAO)")
    print("=" * 65)

    api_key = GEMINI_API_KEY
    if not api_key:
        api_key = input("\nNhập khóa GEMINI_API_KEY của bạn (hoặc set biến môi trường GEMINI_API_KEY): ").strip()
        if not api_key:
            print("[LỖI] Không có API Key, hủy tiến trình.")
            return

    if not XLSX_PATH.exists():
        print(f"[LỖI] Không tìm thấy {XLSX_PATH}")
        return

    system_rules = ""
    if PROMPT_FILE.exists():
        system_rules = open(PROMPT_FILE, 'r', encoding='utf-8').read()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    untranslated_rows = []
    for row_idx in range(2, ws.max_row + 1):
        r_dict = {}
        for col_idx, h in enumerate(headers, start=1):
            r_dict[h] = ws.cell(row=row_idx, column=col_idx).value
        
        vn_val = str(r_dict.get('vietnamese') or '').strip()
        if not vn_val or vn_val.lower() == 'none':
            r_dict['_row_num'] = row_idx
            untranslated_rows.append(r_dict)

    total_untranslated = len(untranslated_rows)
    print(f"\nTổng số dòng cần dịch: {total_untranslated:,} dòng.")
    if total_untranslated == 0:
        print("[THÔNG BÁO] Toàn bộ file đã được dịch 100%!")
        return

    batch_num = 0
    total_batches = (total_untranslated + BATCH_SIZE - 1) // BATCH_SIZE

    for i in range(0, total_untranslated, BATCH_SIZE):
        batch_num += 1
        chunk = untranslated_rows[i:i + BATCH_SIZE]
        print(f"\n--- [BATCH {batch_num}/{total_batches}] Đang dịch {len(chunk)} dòng (ID {chunk[0]['row_id']} -> {chunk[-1]['row_id']}) ---")

        items = [{"id": r["row_id"], "speaker": r["speaker_clean"], "text": r["original_jp"], "scene": r["scene_type"]} for r in chunk]
        user_prompt = f"Dịch các câu sau sang tiếng Việt (Format JSON mảng đối tượng [{{id, vietnamese}}]):\n{json.dumps(items, ensure_ascii=False, indent=2)}"

        res_data = call_gemini_api(api_key, system_rules, user_prompt)
        if res_data and isinstance(res_data, list):
            trans_map = {int(item.get('id', 0)): str(item.get('vietnamese', '')).strip() for item in res_data if item.get('id')}
            applied_count = 0
            for r in chunk:
                rid = int(r['row_id'])
                if rid in trans_map and trans_map[rid]:
                    ws.cell(row=r['_row_num'], column=8, value=trans_map[rid])
                    applied_count += 1

            wb.save(XLSX_PATH)
            print(f"[OK] Đã lưu {applied_count}/{len(chunk)} dòng vào {XLSX_PATH.name} thành công!")
        else:
            print(f"[CẢNH BÁO] Thất bại ở batch {batch_num}, tạm dừng 5 giây...")
            time.sleep(5)

        time.sleep(1)

    print("\n" + "=" * 65)
    print("  >>> HOÀN TẤT DỊCH TOÀN BỘ FILE EXCEL BẰNG GEMINI API!")
    print("=" * 65)


if __name__ == '__main__':
    run_gemini_translator()
