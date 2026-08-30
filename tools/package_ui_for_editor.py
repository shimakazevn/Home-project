import os, sys, json, struct, zipfile, shutil
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = r'E:\HOME_'
ASAR_PATH = os.path.join(PROJECT_ROOT, 'Game', 'resources', 'app.asar')
UI_PACKAGE_DIR = os.path.join(PROJECT_ROOT, 'UI_TRANSLATION_PACKAGE')
ZIP_OUT_PATH = os.path.join(PROJECT_ROOT, 'UI_Translation_And_Graphics_Package.zip')

# Clean output dir
if os.path.exists(UI_PACKAGE_DIR):
    shutil.rmtree(UI_PACKAGE_DIR)
os.makedirs(UI_PACKAGE_DIR, exist_ok=True)

# 1. TẠO FILE EXCEL UI ĐẦY ĐỦ: ui_export.xlsx
print("1. Đang tạo file Excel giao diện: ui_export.xlsx...")
ui_csv_path = os.path.join(PROJECT_ROOT, 'translation', 'ui_export.csv')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UI_Dich_Giao_Dien"

headers = [
    "ID", "File Script", "Dòng Số", "Loại Giao Diện", "Người Nói/Nhãn",
    "Tiếng Nhật Gốc (Original JP)", "BẢN DỊCH TIẾNG VIỆT (DỊCH VÀO ĐÂY)",
    "Thẻ Tag Giữ Nguyên", "Phân Loại", "Ghi Chú Hướng Dẫn"
]

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

import csv
with open(ui_csv_path, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    next(reader, None) # skip header
    for row in reader:
        if len(row) >= 10:
            ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
    for cell in row:
        cell.font = Font(name="Segoe UI", size=10)
        cell.border = thin_border
        if cell.column == 7: # Target column
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="002060")

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 42
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 35

ui_excel_dest = os.path.join(UI_PACKAGE_DIR, 'UI_Text_To_Translate', 'ui_export.xlsx')
os.makedirs(os.path.dirname(ui_excel_dest), exist_ok=True)
wb.save(ui_excel_dest)
shutil.copy(ui_csv_path, os.path.join(UI_PACKAGE_DIR, 'UI_Text_To_Translate', 'ui_export.csv'))
print(f"  [OK] Đã tạo: {ui_excel_dest}")

# 2. TRÍCH XUẤT HÌNH ẢNH GIAO DIỆN (UI IMAGES) TỪ APP.ASAR
print("2. Đang trích xuất toàn bộ ảnh giao diện / nút bấm từ app.asar...")

with open(ASAR_PATH, 'rb') as f:
    f.seek(4)
    header_size = struct.unpack('<I', f.read(4))[0]
    f.seek(12)
    header_len = struct.unpack('<I', f.read(4))[0]
    header_json = f.read(header_len).decode('utf-8')
    header = json.loads(header_json)
    base_offset = 8 + header_size

def walk_files(tree, prefix=''):
    files = []
    for k, v in tree.get('files', {}).items():
        curr = f'{prefix}/{k}' if prefix else k
        if 'files' in v:
            files.extend(walk_files(v, curr))
        else:
            files.append((curr, v.get('size', 0), int(v.get('offset', 0))))
    return files

all_files = walk_files(header)

# Filter for UI-related images
ui_image_patterns = [
    'data/image/button',
    'data/image/config',
    'data/image/system',
    'data/fgimage/chara/button',
    'data/others/plugin/theme_kopanda_09_2/image',
    'data/others/plugin/button_ex',
    'data/others/plugin/uiparts_set',
    'tyrano/images/system',
    'data/image/title',
    'data/image/hud',
]

ui_files_to_extract = []
for file_info in all_files:
    rel_path, size, offset = file_info
    if rel_path.lower().endswith(('.png', '.jpg', '.gif')):
        # Check if it's a UI folder
        if any(pat in rel_path for pat in ui_image_patterns):
            ui_files_to_extract.append(file_info)
        elif rel_path.startswith('data/image/') and not rel_path.startswith('data/image/H') and not rel_path.startswith('data/image/EVCG') and not rel_path.startswith('data/image/kaisou_'):
            ui_files_to_extract.append(file_info)

print(f"  Tìm thấy {len(ui_files_to_extract)} ảnh UI & nút bấm cần biên tập.")

with open(ASAR_PATH, 'rb') as asar_f:
    for rel_path, size, offset in ui_files_to_extract:
        # Category folder mapping
        if 'button' in rel_path.lower():
            cat = '01_Buttons_Nut_Bam'
        elif 'config' in rel_path.lower():
            cat = '02_Config_Cai_Dat'
        elif 'system' in rel_path.lower():
            cat = '03_System_Hop_Thoai_Menu'
        else:
            cat = '04_Other_UI_Giao_Dien_Khac'
            
        out_file_path = os.path.join(UI_PACKAGE_DIR, 'UI_Images_To_Edit', cat, rel_path.replace('/', os.sep))
        os.makedirs(os.path.dirname(out_file_path), exist_ok=True)
        
        asar_f.seek(base_offset + offset)
        data = asar_f.read(size)
        with open(out_file_path, 'wb') as out_f:
            out_f.write(data)

# 3. TẠO FILE HƯỚNG DẪN CHI TIẾT CHO EDITOR & TRANSLATOR
guide_content = """# HƯỚNG DẪN DỊCH THUẬT & BIÊN TẬP GIAO DIỆN (UI) - HOME_

Gói này được tạo riêng để gửi cho **Editor (Thiết kế hình ảnh)** và **Translator (Dịch thuật giao diện)**.

---

## 📁 CẤU TRÚC GÓI NÀY

```
UI_TRANSLATION_PACKAGE/
├── UI_Text_To_Translate/
│   ├── ui_export.xlsx       <-- File Excel dịch text nút bấm, menu, lựa chọn (323 dòng)
│   └── ui_export.csv        <-- File CSV dự phòng
│
└── UI_Images_To_Edit/       <-- Thư mục ảnh giao diện gốc trích xuất từ game
    ├── 01_Buttons_Nut_Bam/  <-- Các nút bấm (Save, Load, Config, Auto, Skip, Back, Log, v.v.)
    ├── 02_Config_Cai_Dat/   <-- Giao diện âm lượng, tốc độ chữ, cài đặt hệ thống
    ├── 03_System_Hop_Thoai_Menu/ <-- Khung thoại, khung tên, icon menu
    └── 04_Other_UI_Giao_Dien_Khac/ <-- Logo, banner, thanh trạng thái
```

---

## ✍️ 1. HƯỚNG DẪN DÀNH CHO TRANSLATOR (DỊCH TEXT UI)

1. Mở file `UI_Text_To_Translate/ui_export.xlsx` bằng Microsoft Excel hoặc Google Sheets.
2. Dịch vào cột **G (BẢN DỊCH TIẾNG VIỆT - Ô MÀU VÀNG)**.
3. **Quy tắc quan trọng:**
   - Dịch ngắn gọn, xúc tích, chuẩn thuật ngữ game Visual Novel (VD: `セーブ` -> `Lưu game`, `ロード` -> `Tải game`, `スキップ` -> `Tua nhanh`, `オート` -> `Tự động`).
   - Nếu gặp biến số như `[font color="yellow"]...[resetfont]`, hãy giữ nguyên thẻ và chỉ dịch phần chữ bên trong.
4. Sau khi dịch xong, lưu lại file và gửi lại cho Trưởng dự án.

---

## 🎨 2. HƯỚNG DẪN DÀNH CHO IMAGE EDITOR (PHOTOSHOP / CANVA)

1. Duyệt các ảnh trong thư mục `UI_Images_To_Edit/`. Ảnh nào có chứa chữ tiếng Nhật cần dịch thì mở bằng Photoshop để edit.
2. **Quy tắc kỹ thuật bắt buộc:**
   - **Giữ nguyên 100% kích thước (Dimensions Pixel Width x Height)** của ảnh gốc (Không scale to hay nhỏ hơn).
   - **Giữ nguyên định dạng `.png` trong suốt (Transparent Alpha)** hoặc `.jpg` như file gốc.
   - **Giữ nguyên tên file và đường dẫn thư mục**.
3. **Font chữ khuyên dùng khi gõ tiếng Việt:**
   - Font Sans-serif bo tròn nhẹ hoặc font hiện đại: **Noto Sans**, **Comfortaa**, **Montserrat**, **Roboto Bold**.
   - Căn giữa (Center alignment) đẹp mắt trên các nút bấm.

---

## 🚀 3. CÁCH ÁP DỤNG LẠI VÀO GAME SAU KHI EDIT XONG

- **Đối với file Text (`ui_export.xlsx` / `ui_export.csv`):**
  Chép đè vào thư mục `translation/` của dự án.
- **Đối với hình ảnh đã edit:**
  Chép các ảnh đã việt hóa vào thư mục `patch/` theo đúng cây thư mục tương ứng (ví dụ: `patch/data/image/button/...`).
- Chạy lệnh build:
  ```bash
  python tools/build_vietnamese_game.py
  ```
  Game sẽ tự động nạp toàn bộ UI text và ảnh mới vào game!
"""

with open(os.path.join(UI_PACKAGE_DIR, 'README_HUONG_DAN_UI.md'), 'w', encoding='utf-8') as f:
    f.write(guide_content)

# 4. NÉN THÀNH FILE .ZIP HOÀN CHỈNH
print(f"3. Đang nén toàn bộ thành file ZIP: {ZIP_OUT_PATH}...")
with zipfile.ZipFile(ZIP_OUT_PATH, 'w', zipfile.ZIP_DEFLATED) as zipf:
    for root, dirs, files in os.walk(UI_PACKAGE_DIR):
        for file in files:
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(full_path, UI_PACKAGE_DIR)
            zipf.write(full_path, rel_path)

zip_size_mb = os.path.getsize(ZIP_OUT_PATH) / (1024 * 1024)
print(f"\n>>> HOÀN TẤT ĐÓNG GÓI UI TRANSLATION PACKAGE!")
print(f"    File ZIP: {ZIP_OUT_PATH} ({zip_size_mb:.2f} MB)")
