import os, sys, json, struct, zipfile, shutil
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = r'E:\HOME_'
UI_PACKAGE_DIR = os.path.join(PROJECT_ROOT, 'UI_TRANSLATION_PACKAGE')
ZIP_OUT_PATH = os.path.join(PROJECT_ROOT, 'UI_Translation_And_Graphics_Package.zip')

# Danh bạ dịch thuật toàn diện cho từng ảnh UI
# (Đã tối ưu Tiếng Việt + Tiếng Anh ngắn gọn cho nút hẹp)
UI_IMAGE_DB = [
    # --- 1. TITLE MENU ---
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_newgame_off.png / title_newgame_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_newgame_off.png",
        "dim": "Nút lớn",
        "jp": "はじめから",
        "vn": "Chơi Mới",
        "en": "New Game",
        "priority": "Bắt buộc sửa",
        "notes": "Chữ New Game hoặc Chơi Mới. Giữ hiệu ứng sáng bóng khi di chuột (_on)."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_load_off.png / title_load_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_load_off.png",
        "dim": "Nút lớn",
        "jp": "つづきから",
        "vn": "Tiếp Tục",
        "en": "Load Game",
        "priority": "Bắt buộc sửa",
        "notes": "Nút tải lại file đã lưu."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_replay_off.png / title_replay_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_replay_off.png",
        "dim": "Nút lớn",
        "jp": "回想",
        "vn": "Xem Lại",
        "en": "Replay",
        "priority": "Bắt buộc sửa",
        "notes": "Vào phòng xem lại các sự kiện."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_Hreplay_off.png / title_Hreplay_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_Hreplay_off.png",
        "dim": "Nút lớn",
        "jp": "H回想",
        "vn": "Cảnh 18+",
        "en": "H-Scenes",
        "priority": "Bắt buộc sửa",
        "notes": "Xem lại các phân cảnh H-Scene."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_EVreplay_off.png / title_EVreplay_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_EVreplay_off.png",
        "dim": "Nút lớn",
        "jp": "イベント回想",
        "vn": "Sự Kiện",
        "en": "Events",
        "priority": "Bắt buộc sửa",
        "notes": "Xem lại hoạt cảnh cốt truyện đặc biệt."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_VoyeurGallery_off.png / title_VoyeurGallery_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_VoyeurGallery_off.png",
        "dim": "Nút lớn",
        "jp": "盗撮ギャラリー",
        "vn": "Xem Trộm",
        "en": "Voyeur CG",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng trưng bày ảnh chụp trộm/theo dõi."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_zenkaihou_off.png / title_zenkaihou_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_zenkaihou_off.png",
        "dim": "Nút lớn",
        "jp": "全開放",
        "vn": "Mở Khóa Hết",
        "en": "Unlock All",
        "priority": "Bắt buộc sửa",
        "notes": "Nút mở khóa toàn bộ CG và Replay cho người chơi."
    },
    {
        "cat": "Title Menu (Màn hình chính)",
        "file": "title_gazohenkou_off.png / title_gazohenkou_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/title_gazohenkou_off.png",
        "dim": "Nút lớn",
        "jp": "画像変更",
        "vn": "Đổi Hình Nền",
        "en": "Change BG",
        "priority": "Khuyên dùng",
        "notes": "Thay đổi hình nền ở Menu Title."
    },

    # --- 2. IN-GAME DAILY ACTIVITY BUTTONS ---
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_con.png / b_job_con_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_con.png",
        "dim": "106x107px",
        "jp": "コンビニ",
        "vn": "Tiện Lợi",
        "en": "Mart",
        "priority": "Bắt buộc sửa",
        "notes": "Làm thêm ở Cửa hàng tiện lợi. Dùng từ 'Tiện Lợi' hoặc 'Mart' cho vừa vặn icon."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_drug.png / b_job_drug_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_drug.png",
        "dim": "106x107px",
        "jp": "ドラッグストア",
        "vn": "Hiệu Thuốc",
        "en": "Pharmacy",
        "priority": "Bắt buộc sửa",
        "notes": "Làm thêm ở Nhà thuốc."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_massa.png / b_job_massa_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_massa.png",
        "dim": "106x107px",
        "jp": "マッサージ",
        "vn": "Massage",
        "en": "Massage",
        "priority": "Bắt buộc sửa",
        "notes": "Làm việc tại tiệm Massage của Rinko."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_hospital.png / b_job_hospital_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_hospital.png",
        "dim": "106x107px",
        "jp": "治験・病院",
        "vn": "Thử Nghiệm",
        "en": "Hospital",
        "priority": "Bắt buộc sửa",
        "notes": "Làm thử nghiệm thuốc ở bệnh viện kiếm tiền."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_fx.png / b_job_fx_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_fx.png",
        "dim": "106x107px",
        "jp": "FX取引",
        "vn": "Giao Dịch FX",
        "en": "FX Trade",
        "priority": "Bắt buộc sửa",
        "notes": "Đầu tư tài chính FX."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_training.png / b_job_training_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_training.png",
        "dim": "106x107px",
        "jp": "筋トレ",
        "vn": "Tập Luyện",
        "en": "Workout",
        "priority": "Bắt buộc sửa",
        "notes": "Tập thể hình tăng thể lực."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_sinnyu.png / b_job_sinnyu_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_sinnyu.png",
        "dim": "106x107px",
        "jp": "侵入",
        "vn": "Đột Nhập",
        "en": "Infiltrate",
        "priority": "Bắt buộc sửa",
        "notes": "Đột nhập vào phòng các nữ chính."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_soudan.png / b_job_soudan_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_soudan.png",
        "dim": "106x107px",
        "jp": "相談",
        "vn": "Tâm Sự",
        "en": "Talk",
        "priority": "Bắt buộc sửa",
        "notes": "Nói chuyện tâm sự với nữ chính."
    },
    {
        "cat": "Daily Job (Nút hoạt động ngày)",
        "file": "b_job_neru.png / b_job_neru_on.png",
        "path": "01_Buttons_Nut_Bam/data/fgimage/chara/button/b_job_neru.png",
        "dim": "106x107px",
        "jp": "寝る",
        "vn": "Đi Ngủ",
        "en": "Sleep",
        "priority": "Bắt buộc sửa",
        "notes": "Kết thúc ca ngày để chuyển sang buổi tối."
    },

    # --- 3. INFILTRATION ACTIONS (ĐỘT NHẬP) ---
    {
        "cat": "Infiltration (Đột nhập)",
        "file": "shinnyu_command_busshoku.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/shinnyu_command_busshoku.png",
        "dim": "Nút chữ nhật",
        "jp": "物色する",
        "vn": "Lục Lọi",
        "en": "Search",
        "priority": "Bắt buộc sửa",
        "notes": "Tìm kiếm đồ vật trong phòng."
    },
    {
        "cat": "Infiltration (Đột nhập)",
        "file": "shinnyu_command_haichi.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/shinnyu_command_haichi.png",
        "dim": "Nút chữ nhật",
        "jp": "道具を配置",
        "vn": "Đặt Dụng Cụ",
        "en": "Place Item",
        "priority": "Bắt buộc sửa",
        "notes": "Đặt thuốc kích dục hoặc camera quay lén."
    },
    {
        "cat": "Infiltration (Đột nhập)",
        "file": "shinnyu_command_kitaku.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/shinnyu_command_kitaku.png",
        "dim": "Nút chữ nhật",
        "jp": "帰宅する",
        "vn": "Rút Lui",
        "en": "Leave",
        "priority": "Bắt buộc sửa",
        "notes": "Rút lui về phòng mình an toàn."
    },

    # --- 4. MAP NAVIGATION (DI CHUYỂN PHÒNG) ---
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_ribingu.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_ribingu.png",
        "dim": "Thanh điều hướng",
        "jp": "リビング",
        "vn": "Phòng Khách",
        "en": "Living Room",
        "priority": "Bắt buộc sửa",
        "notes": "Di chuyển ra phòng khách nhà Naruse."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_nagiroom.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_nagiroom.png",
        "dim": "Thanh điều hướng",
        "jp": "凪の部屋",
        "vn": "Phòng Nagi",
        "en": "Nagi's Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng ngủ của Nagi."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_wasitu.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_wasitu.png",
        "dim": "Thanh điều hướng",
        "jp": "和室",
        "vn": "Phòng Kiểu Nhật",
        "en": "Japanese Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng ngủ của mẹ Rinko."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_datuijyo.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_datuijyo.png",
        "dim": "Thanh điều hướng",
        "jp": "脱衣所",
        "vn": "Phòng Thay Đồ",
        "en": "Dressing Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng thay quần áo giặt."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_huro.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_huro.png",
        "dim": "Thanh điều hướng",
        "jp": "お風呂",
        "vn": "Phòng Tắm",
        "en": "Bathroom",
        "priority": "Bắt buộc sửa",
        "notes": "Bồn tắm nước nóng."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_toire.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_toire.png",
        "dim": "Thanh điều hướng",
        "jp": "トイレ",
        "vn": "Nhà Vệ Sinh",
        "en": "Toilet",
        "priority": "Bắt buộc sửa",
        "notes": "Khu vệ sinh."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_toire_ona.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_toire_ona.png",
        "dim": "Thanh điều hướng",
        "jp": "トイレ（オナニー）",
        "vn": "Tự Sướng",
        "en": "Masturbate",
        "priority": "Bắt buộc sửa",
        "notes": "Tự xử trong nhà vệ sinh để giải tỏa stress."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_byouin_rokka.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_byouin_rokka.png",
        "dim": "Thanh điều hướng",
        "jp": "ロッカールーム",
        "vn": "Phòng Đổi Đồ",
        "en": "Locker Room",
        "priority": "Bắt buộc sửa",
        "notes": "Tủ đồ bác sĩ/y tá ở bệnh viện."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_rihure_sekkyaku.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_rihure_sekkyaku.png",
        "dim": "Thanh điều hướng",
        "jp": "接客室",
        "vn": "Phòng Khách Maid",
        "en": "Service Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng phục vụ khách quán Maid của Tsubomi."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_rihure_hikae.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_rihure_hikae.png",
        "dim": "Thanh điều hướng",
        "jp": "控え室",
        "vn": "Phòng Nghỉ Nhân Viên",
        "en": "Staff Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng nghỉ của nhân viên quán Maid."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_rihure_sityaku.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_rihure_sityaku.png",
        "dim": "Thanh điều hướng",
        "jp": "試着室",
        "vn": "Phòng Thử Đồ",
        "en": "Fitting Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng thay đồng phục Maid."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_massa_sejyutu.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_massa_sejyutu.png",
        "dim": "Thanh điều hướng",
        "jp": "施術室",
        "vn": "Phòng Massage",
        "en": "Massage Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng trị liệu của tiệm Rinko."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_massa_syawa.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_massa_syawa.png",
        "dim": "Thanh điều hướng",
        "jp": "シャワー室",
        "vn": "Phòng Tắm Sen",
        "en": "Shower Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng tắm vòi sen."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_massa_pauda.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_massa_pauda.png",
        "dim": "Thanh điều hướng",
        "jp": "パウダールーム",
        "vn": "Phòng Trang Điểm",
        "en": "Powder Room",
        "priority": "Bắt buộc sửa",
        "notes": "Phòng dặm phấn/trang điểm."
    },
    {
        "cat": "Navigation (Di chuyển địa điểm)",
        "file": "idou_sinnyu_modoru.png / _on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/idou_sinnyu_modoru.png",
        "dim": "Thanh điều hướng",
        "jp": "戻る",
        "vn": "Quay Lại",
        "en": "Back",
        "priority": "Bắt buộc sửa",
        "notes": "Nút quay lại bản đồ trước."
    },

    # --- 5. RELATIONSHIP & MOOD CARDS ---
    {
        "cat": "Mood & Relation (Thẻ quan hệ)",
        "file": "komyu_jyun.png / _2.png / _2_2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/komyu_jyun.png",
        "dim": "Thẻ trạng thái",
        "jp": "純愛",
        "vn": "Thuần Ái",
        "en": "Pure Love",
        "priority": "Bắt buộc sửa",
        "notes": "Trạng thái tình yêu trong sáng."
    },
    {
        "cat": "Mood & Relation (Thẻ quan hệ)",
        "file": "komyu_suki.png / _2.png / _2_2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/komyu_suki.png",
        "dim": "Thẻ trạng thái",
        "jp": "好意",
        "vn": "Yêu Thích",
        "en": "Affection",
        "priority": "Bắt buộc sửa",
        "notes": "Mức độ cảm tình tốt."
    },
    {
        "cat": "Mood & Relation (Thẻ quan hệ)",
        "file": "komyu_kirai.png / _2.png / _2_2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/komyu_kirai.png",
        "dim": "Thẻ trạng thái",
        "jp": "嫌悪",
        "vn": "Chán Ghét",
        "en": "Dislike",
        "priority": "Bắt buộc sửa",
        "notes": "Mức độ thù ghét/bực bội."
    },
    {
        "cat": "Mood & Relation (Thẻ quan hệ)",
        "file": "komyu_sihai.png / _2.png / _2_2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/komyu_sihai.png",
        "dim": "Thẻ trạng thái",
        "jp": "支配 / 狂愛",
        "vn": "Thống Trị",
        "en": "Dominance",
        "priority": "Bắt buộc sửa",
        "notes": "Trạng thái tha hóa/bị kiểm soát hoàn toàn."
    },
    {
        "cat": "Mood & Relation (Thẻ quan hệ)",
        "file": "komyu_end.png / komyu_end2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/komyu_end.png",
        "dim": "Thẻ trạng thái",
        "jp": "終了",
        "vn": "Kết Thúc",
        "en": "Finish",
        "priority": "Bắt buộc sửa",
        "notes": "Kết thúc buổi giao tiếp."
    },

    # --- 6. SYSTEM QUICK TOOLBAR ---
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_save.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_save.png",
        "dim": "Nút thanh menu",
        "jp": "SAVE",
        "vn": "Lưu Game",
        "en": "SAVE",
        "priority": "Tùy chọn (Đã là Tiếng Anh)",
        "notes": "Có thể giữ nguyên chữ tiếng Anh 'SAVE' hoặc đổi thành 'LƯU'."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_load.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_load.png",
        "dim": "Nút thanh menu",
        "jp": "LOAD",
        "vn": "Tải Game",
        "en": "LOAD",
        "priority": "Tùy chọn (Đã là Tiếng Anh)",
        "notes": "Giữ nguyên 'LOAD' hoặc đổi thành 'TẢI'."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_log.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_log.png",
        "dim": "Nút thanh menu",
        "jp": "LOG",
        "vn": "Nhật Ký",
        "en": "LOG",
        "priority": "Tùy chọn",
        "notes": "Xem lại các câu thoại trước đó."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_skip.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_skip.png",
        "dim": "Nút thanh menu",
        "jp": "SKIP",
        "vn": "Tua Nhanh",
        "en": "SKIP",
        "priority": "Tùy chọn",
        "notes": "Tua nhanh qua đoạn thoại."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_auto.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_auto.png",
        "dim": "Nút thanh menu",
        "jp": "AUTO",
        "vn": "Tự Động",
        "en": "AUTO",
        "priority": "Tùy chọn",
        "notes": "Tự động chạy lời thoại."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "menu_btn_config.png / _on.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/menu_btn_config.png",
        "dim": "Nút thanh menu",
        "jp": "CONFIG",
        "vn": "Cài Đặt",
        "en": "CONFIG",
        "priority": "Tùy chọn",
        "notes": "Mở bảng tùy chỉnh âm thanh & tốc độ chữ."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "c_btn_back.png / c_btn_back2.png",
        "path": "01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/c_btn_back.png",
        "dim": "Nút tròn/chữ nhật",
        "jp": "戻る",
        "vn": "Quay Lại",
        "en": "BACK",
        "priority": "Bắt buộc sửa",
        "notes": "Nút quay lại menu chính/game."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "config_reset_off.png / _on.png",
        "path": "02_Config_Cai_Dat/data/others/plugin/theme_kopanda_09_2/image/config/config_reset_off.png",
        "dim": "Nút nhỏ",
        "jp": "初期化",
        "vn": "Mặc Định",
        "en": "RESET",
        "priority": "Bắt buộc sửa",
        "notes": "Khôi phục cài đặt gốc."
    },
    {
        "cat": "System Toolbar (Nút công cụ game)",
        "file": "kettei_off.png / kettei_on.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/kettei_off.png",
        "dim": "Nút chữ nhật",
        "jp": "決定",
        "vn": "Xác Nhận",
        "en": "OK / Confirm",
        "priority": "Bắt buộc sửa",
        "notes": "Nút xác nhận đặt tên nhân vật."
    },

    # --- 7. NIGHT CHOICE (HOẠT ĐỘNG TỐI) ---
    {
        "cat": "Night Choice (Hoạt động ban đêm)",
        "file": "yoru_ie1.png / yoru_ie2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/yoru_ie1.png",
        "dim": "Nút lớn",
        "jp": "家で過ごす",
        "vn": "Ở Nhà",
        "en": "Stay Home",
        "priority": "Bắt buộc sửa",
        "notes": "Nghỉ ngơi buổi tối tại nhà."
    },
    {
        "cat": "Night Choice (Hoạt động ban đêm)",
        "file": "yoru_gaisyutu1.png / yoru_gaisyutu2.png",
        "path": "04_Other_UI_Giao_Dien_Khac/data/image/yoru_gaisyutu1.png",
        "dim": "Nút lớn",
        "jp": "夜の外出",
        "vn": "Ra Ngoài",
        "en": "Go Out",
        "priority": "Bắt buộc sửa",
        "notes": "Đi dạo ban đêm để gặp các nhân vật khác."
    }
]

# 1. TẠO FILE EXCEL HƯỚNG DẪN DỊCH ẢNH UI CHO EDITOR
excel_guide_path = os.path.join(UI_PACKAGE_DIR, 'UI_IMAGE_TRANSLATION_GUIDE.xlsx')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Huong_Dan_Dich_Anh_UI"

headers = [
    "Nhóm Giao Diện",
    "Tên File Ảnh",
    "Đường Dẫn Trong Package",
    "Kích Thước",
    "Tiếng Nhật Gốc",
    "BẢN DỊCH TIẾNG VIỆT (Khuyên dùng)",
    "TIẾNG ANH NGẮN GỌN (Dành cho nút hẹp)",
    "Độ Ưu Tiên",
    "Ghi Chú Đồ Họa Cho Editor"
]

header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for item in UI_IMAGE_DB:
    row_data = [
        item["cat"],
        item["file"],
        item["path"],
        item["dim"],
        item["jp"],
        item["vn"],
        item["en"],
        item["priority"],
        item["notes"]
    ]
    ws.append(row_data)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
    for cell in row:
        cell.font = Font(name="Segoe UI", size=10)
        cell.border = thin_border
        if cell.column == 6: # Tiếng Việt
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="375623")
        elif cell.column == 7: # Tiếng Anh ngắn
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="7F6000")
        elif cell.column == 8 and cell.value == "Bắt buộc sửa":
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")

ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 28
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 45

wb.save(excel_guide_path)
print(f"[OK] Đã tạo Excel hướng dẫn chi tiết cho Editor: {excel_guide_path}")

# 2. TẠO FILE MARKDOWN HƯỚNG DẪN NHANH (README_EDITOR_CHEATSHEET.md)
md_guide_path = os.path.join(UI_PACKAGE_DIR, 'README_EDITOR_CHEATSHEET.md')
with open(md_guide_path, 'w', encoding='utf-8') as f:
    f.write("""# 🎨 BẢNG TRA CỨU DỊCH ẢNH GIAO DIỆN (UI CHEATSHEET FOR EDITOR)

Tài liệu này được biên soạn riêng cho **Editor thiết kế đồ họa** (người không biết tiếng Nhật).
> 💡 **Quy tắc độ dài chữ:** Nếu từ tiếng Việt quá dài khi đặt vào nút nhỏ, bạn hãy dùng **cột Tiếng Anh ngắn gọn** thay thế để đảm bảo tính thẩm mỹ cao nhất!

---

## 1. 🔘 NÚT HOẠT ĐỘNG BAN NGÀY (DAILY JOBS - 106x107px)
Thư mục: `UI_Images_To_Edit/01_Buttons_Nut_Bam/data/fgimage/chara/button/`

| Tên File Ảnh | Tiếng Nhật Gốc | Bản Dịch Tiếng Việt | Tiếng Anh Ngắn Gọn (Khuyên dùng) | Ghi Chú Đồ Họa |
|---|---|---|---|---|
| `b_job_con.png` / `_on.png` | コンビニ | **Tiện Lợi** | **Mart** | Làm thêm ở cửa hàng tiện lợi |
| `b_job_drug.png` / `_on.png` | ドラッグストア | **Hiệu Thuốc** | **Pharmacy** | Làm thêm ở nhà thuốc |
| `b_job_massa.png` / `_on.png` | マッサージ | **Massage** | **Massage** | Làm việc tại tiệm massage |
| `b_job_hospital.png` / `_on.png` | 治験・病院 | **Thử Nghiệm** | **Hospital** | Thử nghiệm thuốc ở bệnh viện |
| `b_job_fx.png` / `_on.png` | FX取引 | **Giao Dịch FX** | **FX Trade** | Đầu tư tài chính FX |
| `b_job_training.png` / `_on.png` | 筋トレ | **Tập Luyện** | **Workout** | Tập thể hình tăng thể lực |
| `b_job_sinnyu.png` / `_on.png` | 侵入 | **Đột Nhập** | **Infiltrate** | Đột nhập vào phòng nữ chính |
| `b_job_soudan.png` / `_on.png` | 相談 | **Tâm Sự** | **Talk** | Nói chuyện tâm sự |
| `b_job_neru.png` / `_on.png` | 寝る | **Đi Ngủ** | **Sleep** | Kết thúc ngày |

---

## 2. 🎮 MENU CHÍNH (TITLE MENU)
Thư mục: `UI_Images_To_Edit/04_Other_UI_Giao_Dien_Khac/data/image/`

| Tên File Ảnh | Tiếng Nhật Gốc | Bản Dịch Tiếng Việt | Tiếng Anh Ngắn Gọn | Ghi Chú |
|---|---|---|---|---|
| `title_newgame_off.png` / `_on.png` | はじめから | **Chơi Mới** | **New Game** | Bắt đầu game từ đầu |
| `title_load_off.png` / `_on.png` | つづきから | **Tiếp Tục** | **Load Game** | Tải lại file đã lưu |
| `title_replay_off.png` / `_on.png` | 回想 | **Xem Lại** | **Replay** | Xem lại cốt truyện |
| `title_Hreplay_off.png` / `_on.png` | H回想 | **Cảnh 18+** | **H-Scenes** | Xem lại cảnh người lớn |
| `title_EVreplay_off.png` / `_on.png` | イベント回想 | **Sự Kiện** | **Events** | Xem lại sự kiện đặc biệt |
| `title_VoyeurGallery_off.png` / `_on.png` | 盗撮ギャラリー | **Xem Trộm** | **Voyeur CG** | Phòng ảnh chụp trộm |
| `title_zenkaihou_off.png` / `_on.png` | 全開放 | **Mở Khóa Hết** | **Unlock All** | Mở khóa toàn bộ CG |
| `title_gazohenkou_off.png` / `_on.png` | 画像変更 | **Đổi Nền** | **Change BG** | Thay đổi ảnh nền Title |

---

## 3. 🚪 THANH ĐIỀU HƯỚNG BẢN ĐỒ (NAVIGATION BARS)
Thư mục: `UI_Images_To_Edit/04_Other_UI_Giao_Dien_Khac/data/image/`

| Tên File Ảnh | Tiếng Nhật Gốc | Bản Dịch Tiếng Việt | Tiếng Anh Ngắn Gọn |
|---|---|---|---|
| `idou_ribingu.png` / `_on.png` | リビング | **Phòng Khách** | **Living Room** |
| `idou_nagiroom.png` / `_on.png` | 凪の部屋 | **Phòng Nagi** | **Nagi's Room** |
| `idou_wasitu.png` / `_on.png` | 和室 | **Phòng Nhật** | **Japanese Room** |
| `idou_datuijyo.png` / `_on.png` | 脱衣所 | **Phòng Thay Đồ** | **Dressing Room** |
| `idou_huro.png` / `_on.png` | お風呂 | **Phòng Tắm** | **Bathroom** |
| `idou_toire.png` / `_on.png` | トイレ | **Nhà Vệ Sinh** | **Toilet** |
| `idou_toire_ona.png` | トイレ（オナニー） | **Tự Sướng** | **Masturbate** |
| `idou_byouin_rokka.png` / `_on.png` | ロッカールーム | **Tủ Đồ** | **Locker Room** |
| `idou_rihure_sekkyaku.png` / `_on.png` | 接客室 | **Phòng Khách** | **Service Room** |
| `idou_rihure_hikae.png` / `_on.png` | 控え室 | **Phòng Nghỉ** | **Staff Room** |
| `idou_rihure_sityaku.png` / `_on.png` | 試着室 | **Phòng Thử Đồ** | **Fitting Room** |
| `idou_massa_sejyutu.png` / `_on.png` | 施術室 | **Phòng Trị Liệu** | **Massage Room** |
| `idou_massa_syawa.png` / `_on.png` | シャワー室 | **Phòng Tắm Sen** | **Shower** |
| `idou_massa_pauda.png` / `_on.png` | パウダールーム | **Trang Điểm** | **Powder Room** |
| `idou_sinnyu_modoru.png` / `_on.png` | 戻る | **Quay Lại** | **Back** |

---

## 4. 🗃️ CÔNG CỤ HỆ THỐNG (SYSTEM BUTTONS)
Thư mục: `UI_Images_To_Edit/01_Buttons_Nut_Bam/data/others/plugin/theme_kopanda_09_2/image/button/`

| Tên File Ảnh | Tiếng Nhật Gốc | Bản Dịch Tiếng Việt / Tiếng Anh Khuyên Dùng |
|---|---|---|
| `c_btn_back.png` / `c_btn_back2.png` | 戻る | **BACK / Quay Lại** |
| `menu_btn_save.png` / `_on.png` | SAVE | **SAVE / Lưu** |
| `menu_btn_load.png` / `_on.png` | LOAD | **LOAD / Tải** |
| `menu_btn_log.png` / `_on.png` | LOG | **LOG / Nhật Ký** |
| `menu_btn_skip.png` / `_on.png` | SKIP | **SKIP / Tua** |
| `menu_btn_auto.png` / `_on.png` | AUTO | **AUTO / Tự Động** |
| `menu_btn_config.png` / `_on.png` | CONFIG | **CONFIG / Cài Đặt** |
| `config_reset_off.png` / `_on.png` | 初期化 | **RESET / Mặc Định** |
| `kettei_off.png` / `kettei_on.png` | 決定 | **OK / Xác Nhận** |

---

## 5. 🌙 LỰA CHỌN BUỔI TỐI & ĐỘT NHẬP
Thư mục: `UI_Images_To_Edit/04_Other_UI_Giao_Dien_Khac/data/image/`

| Tên File Ảnh | Tiếng Nhật Gốc | Bản Dịch Tiếng Việt | Tiếng Anh Ngắn |
|---|---|---|---|
| `yoru_ie1.png` / `yoru_ie2.png` | 家で過ごす | **Ở Nhà** | **Stay Home** |
| `yoru_gaisyutu1.png` / `yoru_gaisyutu2.png` | 夜の外出 | **Ra Ngoài** | **Go Out** |
| `shinnyu_command_busshoku.png` / `_on.png` | 物色する | **Lục Lọi** | **Search** |
| `shinnyu_command_haichi.png` / `_on.png` | 道具を配置 | **Đặt Đồ** | **Place Item** |
| `shinnyu_command_kitaku.png` / `_on.png` | 帰宅する | **Rút Lui** | **Leave** |

---

### 📌 QUY TẮC BẮT BUỘC KHI XUẤT ẢNH:
1. **Kích thước ảnh:** Giữ nguyên 100% kích thước pixel gốc (Width x Height).
2. **Định dạng file:** Xuất dạng `.png` trong suốt (Transparent background), không nền trắng.
3. **Font chữ đẹp:** Dùng **Noto Sans**, **Montserrat**, **Roboto Bold** hoặc **Comfortaa**.
""")

# 3. NÉN LẠI THÀNH GÓI ZIP HOÀN CHỈNH
print(f"Đang nén lại thành file ZIP hoàn chỉnh: {ZIP_OUT_PATH}...")
with zipfile.ZipFile(ZIP_OUT_PATH, 'w', zipfile.ZIP_DEFLATED) as zipf:
    for root, dirs, files in os.walk(UI_PACKAGE_DIR):
        for file in files:
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(full_path, UI_PACKAGE_DIR)
            zipf.write(full_path, rel_path)

zip_size_mb = os.path.getsize(ZIP_OUT_PATH) / (1024 * 1024)
print(f"\n>>> HOÀN TẤT ĐÓNG GÓI UI TRANSLATION PACKAGE (CÓ BẢNG DỊCH CHO EDITOR)!")
print(f"    File ZIP: {ZIP_OUT_PATH} ({zip_size_mb:.2f} MB)")
